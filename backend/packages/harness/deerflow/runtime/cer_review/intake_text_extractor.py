"""CER Raw Project Intake — Text Extractor

Deterministic text extraction from PDF, DOCX, XLSX, TXT using established libraries.
NO LLM, NO AGENT. Pure Python wrappers.

Frozen baseline: CER_RAW_PROJECT_INTAKE_AGENT_VS_PROGRAM_BOUNDARY.md
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ── Text Extraction ─────────────────────────────────────────────────────────────


class TextExtractionError(Exception):
    """Raised when text extraction fails."""
    pass


def extract_text(file_path: Path) -> str:
    """Extract text from a file based on its extension.

    Supports: PDF, DOCX, XLSX, TXT, MD
    Raises TextExtractionError on failure.
    """
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    elif ext in {".txt", ".md"}:
        return _extract_txt(file_path)
    else:
        raise TextExtractionError(f"Unsupported file type: {ext}")


def extract_text_batch(
    file_paths: list[Path],
    output_dir: Path,
) -> dict[str, Any]:
    """Extract text from multiple files and write to output directory.

    Returns a document_text_index.json-compatible dict.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    results = []
    total_extracted = 0
    total_failed = 0

    for file_path in file_paths:
        file_id = _path_to_file_id(file_path)
        try:
            text = extract_text(file_path)
            output_name = f"{file_id}_{file_path.stem}.txt"
            output_path = output_dir / output_name
            output_path.write_text(text, encoding="utf-8")
            results.append({
                "file_id": file_id,
                "relative_path": str(file_path),
                "extractable": True,
                "extracted": True,
                "text_extracted_path": str(output_path.relative_to(output_dir.parent)),
                "extraction_method": _extraction_method(file_path.suffix),
                "character_count": len(text),
                "failure_reason": None,
            })
            total_extracted += 1
        except TextExtractionError as e:
            results.append({
                "file_id": file_id,
                "relative_path": str(file_path),
                "extractable": True,
                "extracted": False,
                "text_extracted_path": None,
                "extraction_method": _extraction_method(file_path.suffix),
                "character_count": 0,
                "failure_reason": str(e),
            })
            total_failed += 1
        except Exception as e:
            results.append({
                "file_id": file_id,
                "relative_path": str(file_path),
                "extractable": True,
                "extracted": False,
                "text_extracted_path": None,
                "extraction_method": _extraction_method(file_path.suffix),
                "character_count": 0,
                "failure_reason": f"unexpected_error: {e}",
            })
            total_failed += 1

    return {
        "schema_name": "cer_intake_document_text_index",
        "schema_version": "v1",
        "generated_at": _utc_now(),
        "total_files_attempted": len(file_paths),
        "total_files_extracted": total_extracted,
        "total_files_failed": total_failed,
        "files": results,
    }


# ── PDF Extraction ──────────────────────────────────────────────────────────────


def _extract_pdf(file_path: Path) -> str:
    """Extract text from PDF using pdfplumber (preferred) or pypdf fallback."""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            pages = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
    except ImportError:
        pass

    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)
    except ImportError:
        raise TextExtractionError(
            f"Neither pdfplumber nor pypdf available for PDF extraction: {file_path}"
        )


# ── DOCX Extraction ────────────────────────────────────────────────────────────


def _extract_docx(file_path: Path) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        raise TextExtractionError(
            f"python-docx not available for DOCX extraction: {file_path}"
        )


# ── XLSX Extraction ────────────────────────────────────────────────────────────


def _extract_xlsx(file_path: Path) -> str:
    """Extract text from XLSX using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            lines.append(f"[Sheet: {sheet_name}]")
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in row_vals):
                    lines.append(" | ".join(row_vals))
        return "\n".join(lines)
    except ImportError:
        raise TextExtractionError(
            f"openpyxl not available for XLSX extraction: {file_path}"
        )


# ── Plain Text Extraction ───────────────────────────────────────────────────────


def _extract_txt(file_path: Path) -> str:
    """Read plain text file."""
    return file_path.read_text(encoding="utf-8", errors="replace")


# ── Helpers ────────────────────────────────────────────────────────────────────


def _extraction_method(ext: str) -> str:
    method_map = {
        ".pdf": "pdfplumber",
        ".docx": "python-docx",
        ".xlsx": "openpyxl",
        ".txt": "txt",
        ".md": "txt",
    }
    return method_map.get(ext.lower(), "unknown")


def _path_to_file_id(path: Path) -> str:
    """Generate a consistent file ID from path."""
    import hashlib
    name = path.name
    h = hashlib.md5(name.encode()).hexdigest()[:6]
    return f"F-{h.upper()}"


def _utc_now() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()
