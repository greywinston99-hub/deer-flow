"""Minimal runner glue for the RMF Review Workflow v1.1 skeleton.

This module intentionally does not pretend DeerFlow has a native workflow
engine for this review chain. It provides a small bridge that:

1. loads ``workflows/rmf_review_v1_1.yaml``
2. maps workflow steps to Python handlers
3. writes artifacts into DeerFlow's per-thread outputs directory
4. supports dry-run and smoke-run for the first five nodes
"""

from __future__ import annotations

import csv
import json
import logging
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import yaml

from deerflow.config.paths import get_paths

logger = logging.getLogger(__name__)

_SUPPORTED_STEP_IDS = (
    "rmf_intake_agent",
    "rmf_parse_normalize_agent",
    "fmea_precheck_agent",
    "rmf_precheck_agent",
    "rmf_dimension_review_agent",
    "rmf_human_boundary_agent",
    "rmf_report_agent",
)

_COLUMN_ALIASES = {
    "risk_id": ("risk id", "risk_id", "id", "risk no", "risk number", "序号"),
    "hazard": ("hazard", "危害"),
    "hazardous_situation": ("hazardous situation", "hazardous_situation", "危险情况"),
    "harm": ("harm", "伤害"),
    "failure_mode": ("failure mode", "failure_mode", "失效模式"),
    "effect": ("effect", "failure effect", "potential effect", "potential effect(s) of failure", "失效影响"),
    "cause": ("cause", "failure cause", "失效原因"),
    "sequence_of_events": ("sequence of events", "sequence_of_events", "事件序列"),
    "probability": ("probability", "occurrence", "p", "发生频度", "发生度"),
    "occurrence": ("occurrence", "probability", "o", "发生频度", "发生度"),
    "severity": ("severity", "s", "严重度"),
    "detection": ("detection", "detectability", "d", "探测度"),
    "risk_level": ("risk level", "risk_level", "initial risk", "residual risk level", "风险等级"),
    "rpn": ("rpn",),
    "controls": ("controls", "risk control", "mitigation", "control measure", "现行预防控制", "风险控制措施"),
    "current_control": ("current control", "current controls", "prevention control", "detection control", "现行预防控制"),
    "verification_evidence": ("verification evidence", "verification", "evidence", "验证证据"),
    "residual_risk": ("residual risk", "residual_risk", "剩余风险"),
    "acceptance_conclusion": ("acceptance conclusion", "acceptability", "risk acceptability", "可接受性结论"),
    "hazard_analysis_ref": ("hazard analysis ref", "ha ref"),
}


@dataclass
class RMFRunResult:
    thread_id: str
    run_id: str
    mode: str
    workflow_name: str
    executed_steps: list[str]
    artifact_root_virtual: str
    artifact_root_actual: str


class RMFReviewRunner:
    """Small bridge from workflow YAML to executable step handlers."""

    def __init__(
        self,
        *,
        repo_root: str | Path,
        workflow_path: str | Path,
        project_profile_path: str | Path,
        input_root: str | Path | None = None,
        thread_id: str | None = None,
    ) -> None:
        self.repo_root = Path(repo_root).resolve()
        self.workflow_path = Path(workflow_path).resolve()
        self.project_profile_path = Path(project_profile_path).resolve()
        self.input_root_override = Path(input_root).resolve() if input_root else None
        self.thread_id = thread_id or self._make_thread_id()

        self.paths = get_paths()
        self.paths.ensure_thread_dirs(self.thread_id)

        self.workflow = self._load_yaml(self.workflow_path)
        self.project_profile = self._load_yaml(self.project_profile_path)

        self.run_id = self._make_run_id()
        self.workflow_name = str(self.workflow.get("workflow_name", "rmf_review_v1_1"))

        artifact_root_template = self.workflow["runtime_defaults"]["artifact_root"]
        self.artifact_root_virtual = self._render_template(str(artifact_root_template))
        self.artifact_root_actual = self.paths.resolve_virtual_path(self.thread_id, self.artifact_root_virtual)
        self.artifact_root_actual.mkdir(parents=True, exist_ok=True)

        self.step_map = {
            "rmf_intake_agent": self._run_intake,
            "rmf_parse_normalize_agent": self._run_parse_normalize,
            "fmea_precheck_agent": self._run_fmea_precheck,
            "rmf_precheck_agent": self._run_rmf_precheck,
            "rmf_dimension_review_agent": self._run_rmf_dimension_review,
            "rmf_human_boundary_agent": self._run_human_boundary,
            "rmf_report_agent": self._run_report,
        }

    def run(self, *, mode: str = "smoke-run") -> RMFRunResult:
        executed_steps: list[str] = []
        self._write_run_context(mode=mode)

        if mode == "dry-run":
            self._write_json(
                self._artifact_path("00_manifest", "dry_run_plan.json"),
                self._build_dry_run_plan(),
            )
            return RMFRunResult(
                thread_id=self.thread_id,
                run_id=self.run_id,
                mode=mode,
                workflow_name=self.workflow_name,
                executed_steps=executed_steps,
                artifact_root_virtual=self.artifact_root_virtual,
                artifact_root_actual=str(self.artifact_root_actual),
            )

        for step in self.workflow.get("ordered_steps", []):
            step_id = step.get("step_id")
            if step_id not in _SUPPORTED_STEP_IDS:
                break
            handler = self.step_map[step_id]
            handler(step)
            executed_steps.append(step_id)

        self._write_json(
            self._artifact_path("00_manifest", "run_summary.json"),
            {
                "thread_id": self.thread_id,
                "run_id": self.run_id,
                "workflow_name": self.workflow_name,
                "mode": mode,
                "executed_steps": executed_steps,
                "artifact_root_virtual": self.artifact_root_virtual,
                "artifact_root_actual": str(self.artifact_root_actual),
            },
        )
        return RMFRunResult(
            thread_id=self.thread_id,
            run_id=self.run_id,
            mode=mode,
            workflow_name=self.workflow_name,
            executed_steps=executed_steps,
            artifact_root_virtual=self.artifact_root_virtual,
            artifact_root_actual=str(self.artifact_root_actual),
        )

    def _run_intake(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("00_manifest")
        prompt_text = self._load_prompt_for_step(step)
        documents = list(self.project_profile.get("input_package", {}).get("documents", []))
        input_root = self._resolve_input_root()
        inventory: list[dict[str, Any]] = []

        for index, doc in enumerate(documents, start=1):
            raw_path = doc.get("path")
            resolved_path = self._resolve_source_path(raw_path, input_root)
            exists = resolved_path.exists()
            status = "present" if exists else "missing"
            entry = {
                "inventory_id": f"doc_{index:03d}",
                "document_id": doc.get("source_ref", {}).get("document_id") or f"document_{index:03d}",
                "doc_type": doc.get("doc_type"),
                "label": doc.get("label"),
                "declared_path": raw_path,
                "resolved_path": str(resolved_path),
                "required_for_p0": bool(doc.get("required_for_p0", False)),
                "blocking_for_p0": bool(doc.get("required_for_p0", False)),
                "status": status,
                "resolution_notes": [],
                "source_ref": {
                    "document_id": doc.get("source_ref", {}).get("document_id") or f"document_{index:03d}",
                    "path": raw_path,
                },
            }
            inventory.append(entry)
        hazard_resolution = self._apply_hazard_analysis_resolution(inventory)
        missing_required = [item for item in inventory if item["blocking_for_p0"] and item["status"] != "present"]

        explicit_fmea_present = any(item["doc_type"] == "FMEA" and item["status"] == "present" for item in inventory)
        explicit_hazard_present = any(item["doc_type"] == "Hazard_Analysis" and item["status"] in {"present", "aliased_to_fmea"} for item in inventory)

        run_manifest = {
            "workflow_name": self.workflow_name,
            "workflow_version": self.workflow.get("workflow_version"),
            "run_id": self.run_id,
            "thread_id": self.thread_id,
            "step_id": "rmf_intake_agent",
            "institution_profile": self.project_profile.get("institution_profile", {}),
            "primary_review_object": self.project_profile.get("primary_review_object"),
            "project_profile_path": str(self.project_profile_path),
            "input_root": str(input_root),
            "artifact_root_virtual": self.artifact_root_virtual,
            "artifact_root_actual": str(self.artifact_root_actual),
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "human_gate_required": self.project_profile.get("review_scope", {}).get("human_gate_required", True),
            "fmea_status": {
                "fmea_present": explicit_fmea_present,
                "hazard_analysis_present": explicit_hazard_present,
                "explicit_review_required": True,
            },
            "hazard_resolution": hazard_resolution,
        }

        missing_lines = [
            "# Missing Items Report",
            "",
            f"- Run ID: `{self.run_id}`",
            f"- Input Root: `{input_root}`",
            f"- Required Missing Count: `{len(missing_required)}`",
            f"- FMEA Present: `{explicit_fmea_present}`",
            f"- Hazard Analysis Present: `{explicit_hazard_present}`",
            "",
        ]
        if missing_required:
            missing_lines.append("## Blocking Missing Items")
            missing_lines.append("")
            for item in missing_required:
                missing_lines.append(
                    f"- `{item['doc_type']}` | `{item['label']}` | `{item['declared_path']}` | status=`{item['status']}`"
                )
        else:
            missing_lines.append("## Blocking Missing Items")
            missing_lines.append("")
            missing_lines.append("- None")

        missing_lines.append("")
        missing_lines.append("## Hazard Analysis Resolution")
        missing_lines.append("")
        missing_lines.append(f"- Mode: `{hazard_resolution['resolution_mode']}`")
        missing_lines.append(f"- Reason: {hazard_resolution['reason']}")
        if hazard_resolution.get("alias_target"):
            missing_lines.append(f"- Alias Target: `{hazard_resolution['alias_target']}`")

        self._write_json(step_dir / "run_manifest.json", run_manifest)
        self._write_json(step_dir / "input_inventory.json", {"documents": inventory, "input_root": str(input_root)})
        self._write_text(step_dir / "missing_items_report.md", "\n".join(missing_lines) + "\n")

    def _run_parse_normalize(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("01_parse")
        prompt_text = self._load_prompt_for_step(step)
        inventory = self._read_json(self._artifact_path("00_manifest", "input_inventory.json"))["documents"]

        rmf_docs = [item for item in inventory if item["doc_type"] == "RMF" and item["status"] == "present"]
        fmea_docs = [item for item in inventory if item["doc_type"] in {"FMEA", "Hazard_Analysis"} and item["status"] == "present"]

        normalized_profile = {
            "project_id": self.project_profile.get("project_id"),
            "project_name": self.project_profile.get("project_name"),
            "institution_profile": self.project_profile.get("institution_profile", {}),
            "primary_review_object": self.project_profile.get("primary_review_object"),
            "input_root": str(self._resolve_input_root()),
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "document_counts": {
                "rmf": len(rmf_docs),
                "fmea_or_hazard": len(fmea_docs),
                "all": len(inventory),
            },
        }

        rmf_normalized = self._build_rmf_normalized(rmf_docs)

        parsed_fmea = self._build_fmea_normalized(fmea_docs)

        cross_doc_entities = {
            "project_id": self.project_profile.get("project_id"),
            "document_entities": [
                {
                    "document_id": item["document_id"],
                    "doc_type": item["doc_type"],
                    "path": item["resolved_path"],
                }
                for item in inventory
            ],
            "risk_ids": [row["risk_id"] for row in parsed_fmea.get("rows", []) if row.get("risk_id")],
        }

        device_name = self.project_profile.get("device_context", {}).get("device_name")
        term_map = {
            "canonical_terms": [
                {"term": "RMF", "meaning": "Risk Management File"},
                {"term": "FMEA", "meaning": "Failure Mode and Effects Analysis"},
                {"term": "Hazard Analysis", "meaning": "Explicit hazard analysis input"},
            ],
            "project_terms": ([{"term": device_name, "meaning": "device_name"}] if device_name else []),
        }

        self._write_json(step_dir / "project_profile.normalized.json", normalized_profile)
        self._write_json(step_dir / "rmf_normalized.json", rmf_normalized)
        self._write_json(step_dir / "fmea_normalized.json", parsed_fmea)
        self._write_json(step_dir / "cross_doc_entities.json", cross_doc_entities)
        self._write_json(step_dir / "term_map.json", term_map)

    def _run_fmea_precheck(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("02_fmea_precheck")
        prompt_text = self._load_prompt_for_step(step)
        fmea_normalized = self._read_json(self._artifact_path("01_parse", "fmea_normalized.json"))
        rows = list(fmea_normalized.get("rows", []))

        duplicate_risk_ids = self._duplicates([row.get("risk_id", "") for row in rows if row.get("risk_id")])
        findings: list[dict[str, Any]] = []
        empty_rows = 0
        orphan_rows = 0

        for idx, row in enumerate(rows, start=1):
            missing_fields: list[str] = []
            if not row.get("risk_id"):
                missing_fields.append("risk_id")
            if not (row.get("probability") or {}).get("raw_value"):
                missing_fields.append("probability")
            if not (row.get("severity") or {}).get("raw_value"):
                missing_fields.append("severity")
            if not ((row.get("risk_level") or {}).get("raw_value") or (row.get("rpn") or {}).get("raw_value")):
                missing_fields.append("risk_level_or_rpn")
            if not row.get("controls"):
                missing_fields.append("controls")
            if not row.get("verification_evidence"):
                missing_fields.append("verification_evidence")
            if not row.get("residual_risk", {}).get("summary"):
                missing_fields.append("residual_risk")
            if not row.get("acceptance_conclusion", {}).get("status"):
                missing_fields.append("acceptance_conclusion")
            if not row.get("source_ref"):
                missing_fields.append("source_ref")

            nonempty_signal_fields = [
                row.get("risk_id"),
                row.get("hazard"),
                row.get("hazardous_situation"),
                row.get("harm"),
                row.get("failure_mode"),
            ]
            if not any(nonempty_signal_fields):
                empty_rows += 1
            if row.get("risk_id") and not row.get("controls"):
                orphan_rows += 1

            if row.get("raw_cells") and row.get("mapping_status") == "raw_only":
                findings.append(
                    {
                        "row_index": idx,
                        "risk_id": row.get("risk_id"),
                        "finding_type": "mapping_incomplete",
                        "missing_fields": ["semantic_mapping"],
                        "source_ref": row.get("source_ref"),
                    }
                )

            if missing_fields:
                findings.append(
                    {
                        "row_index": idx,
                        "risk_id": row.get("risk_id"),
                        "finding_type": "missing_required_fields",
                        "missing_fields": missing_fields,
                        "source_ref": row.get("source_ref"),
                    }
                )

        for duplicated in duplicate_risk_ids:
            findings.append(
                {
                    "finding_type": "duplicate_risk_id",
                    "risk_id": duplicated,
                }
            )

        report = {
            "step_id": "fmea_precheck_agent",
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "document_id": fmea_normalized.get("document_id"),
            "document_type": fmea_normalized.get("document_type"),
            "row_count": len(rows),
            "duplicate_risk_ids": duplicate_risk_ids,
            "empty_row_count": empty_rows,
            "orphan_row_count": orphan_rows,
            "mapping_incomplete_row_count": sum(1 for finding in findings if finding.get("finding_type") == "mapping_incomplete"),
            "manual_review_needed": bool(findings) or not rows,
            "structural_status": self._determine_precheck_status(rows=rows, findings=findings),
            "findings": findings,
        }

        self._write_json(step_dir / "fmea_precheck_report.json", report)
        self._write_text(
            step_dir / "fmea_precheck_report.md",
            self._render_precheck_markdown(report),
        )

    def _run_rmf_precheck(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("03_rmf_precheck")
        prompt_text = self._load_prompt_for_step(step)
        rmf_normalized = self._read_json(self._artifact_path("01_parse", "rmf_normalized.json"))
        term_map = self._read_json(self._artifact_path("01_parse", "term_map.json"))
        cross_doc_entities = self._read_json(self._artifact_path("01_parse", "cross_doc_entities.json"))
        input_inventory = self._read_json(self._artifact_path("00_manifest", "input_inventory.json"))
        run_manifest = self._read_json(self._artifact_path("00_manifest", "run_manifest.json"))

        text_lines = list(rmf_normalized.get("text_lines", []))
        section_hints = list(rmf_normalized.get("section_hints", []))
        findings: list[dict[str, Any]] = []
        evidence_hints: list[dict[str, Any]] = []

        check_specs = [
            (
                "risk_management_plan",
                "风险管理计划/主体存在性检查",
                ["风险管理计划", "risk management plan", "风险管理报告", "risk management report"],
                "Section or title hint for the RMF main body.",
            ),
            (
                "risk_analysis_matrix",
                "风险分析矩阵存在性检查",
                ["风险分析", "风险评价准则", "risk analysis", "probability", "severity"],
                "Matrix-like risk analysis terms were searched.",
            ),
            (
                "risk_level_definition",
                "风险等级定义表/等级线索检查",
                ["严重度", "severity", "发生概率", "概率", "probability", "发生频度"],
                "Risk level definition clues were searched.",
            ),
            (
                "traceability_matrix",
                "可追溯矩阵或可追溯线索检查",
                ["可追溯", "traceability"],
                "Traceability clues were searched.",
            ),
            (
                "residual_risk_evaluation",
                "剩余风险评价存在性检查",
                ["剩余风险", "residual risk", "综合剩余风险"],
                "Residual risk clues were searched.",
            ),
            (
                "production_post_production_information",
                "生产和生产后信息章节/线索检查",
                ["生产和生产后", "生产后信息", "post-production", "production and post-production"],
                "Production/post-production clues were searched.",
            ),
            (
                "iso_14971_reference",
                "ISO 14971 引用线索检查",
                ["iso 14971", "iso14971", "yy/t 0316"],
                "ISO 14971 citation clues were searched.",
            ),
            (
                "risk_control_hierarchy",
                "风险控制三步法线索检查",
                ["固有安全设计", "防护措施", "安全信息", "三步法", "inherent safety", "protective measures", "information for safety"],
                "Three-step risk control hierarchy clues were searched.",
            ),
        ]

        checks: list[dict[str, Any]] = []
        for check_key, label, keywords, rationale in check_specs:
            match = self._search_text_hints(text_lines, keywords)
            state = "present" if match else "weakly_evidenced"
            source_refs = match["source_refs"] if match else rmf_normalized.get("source_refs", [])
            evidence = match["evidence"] if match else []
            if match:
                evidence_hints.append(
                    {
                        "check_key": check_key,
                        "label": label,
                        "evidence": evidence,
                        "source_refs": source_refs,
                    }
                )
            else:
                findings.append(
                    {
                        "finding_type": "weak_evidence",
                        "check_key": check_key,
                        "label": label,
                        "severity": "medium",
                        "detail": rationale,
                        "source_refs": source_refs,
                    }
                )
            checks.append(
                {
                    "check_key": check_key,
                    "label": label,
                    "state": state,
                    "evidence": evidence,
                    "source_refs": source_refs,
                }
            )

        term_consistency = self._build_term_consistency(
            term_map=term_map,
            text_lines=text_lines,
            cross_doc_entities=cross_doc_entities,
            input_inventory=input_inventory,
        )
        if term_consistency["finding"]:
            findings.append(term_consistency["finding"])

        fmea_followup_needed = run_manifest.get("fmea_status", {}).get("explicit_review_required", False)
        if fmea_followup_needed:
            findings.append(
                {
                    "finding_type": "manual_followup_required",
                    "check_key": "fmea_traceability_dependency",
                    "label": "FMEA structural issues require manual RMF traceability follow-up",
                    "severity": "medium",
                    "detail": "RMF precheck completed with evidence hints, but FMEA normalization remains partial and requires reviewer confirmation.",
                    "source_refs": rmf_normalized.get("source_refs", []),
                }
            )

        report = {
            "step_id": "rmf_precheck_agent",
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "document_ids": rmf_normalized.get("document_ids", []),
            "primary_review_object": rmf_normalized.get("primary_review_object"),
            "checks": checks,
            "section_hints": section_hints,
            "evidence_hints": evidence_hints,
            "term_consistency": term_consistency["summary"],
            "manual_review_needed": True,
            "findings": findings,
            "structural_status": self._determine_rmf_precheck_status(checks=checks, findings=findings, evidence_hints=evidence_hints),
        }

        self._write_json(step_dir / "rmf_precheck_report.json", report)
        self._write_text(step_dir / "rmf_precheck_report.md", self._render_rmf_precheck_markdown(report))

    def _run_rmf_dimension_review(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("04_dimension_review")
        prompt_text = self._load_prompt_for_step(step)
        rmf_normalized = self._read_json(self._artifact_path("01_parse", "rmf_normalized.json"))
        fmea_normalized = self._read_json(self._artifact_path("01_parse", "fmea_normalized.json"))
        fmea_precheck = self._read_json(self._artifact_path("02_fmea_precheck", "fmea_precheck_report.json"))
        rmf_precheck = self._read_json(self._artifact_path("03_rmf_precheck", "rmf_precheck_report.json"))
        term_map = self._read_json(self._artifact_path("01_parse", "term_map.json"))
        cross_doc_entities = self._read_json(self._artifact_path("01_parse", "cross_doc_entities.json"))
        input_inventory = self._read_json(self._artifact_path("00_manifest", "input_inventory.json"))
        run_manifest = self._read_json(self._artifact_path("00_manifest", "run_manifest.json"))

        sections_by_key = {item["section_key"]: item for item in rmf_normalized.get("section_hints", [])}
        checks_by_key = {item["check_key"]: item for item in rmf_precheck.get("checks", [])}
        fmea_rows = list(fmea_normalized.get("rows", []))
        fmea_findings = list(fmea_precheck.get("findings", []))
        rmf_findings = list(rmf_precheck.get("findings", []))
        inventory_docs = list(input_inventory.get("documents", []))

        dimensions = {
            "COMP": self._build_dimension_result(
                dimension_id="COMP",
                label="Completeness",
                evidence_hints=self._collect_dimension_evidence(
                    checks_by_key,
                    keys=["risk_management_plan", "risk_analysis_matrix", "risk_level_definition", "residual_risk_evaluation"],
                ),
                findings=self._collect_completeness_findings(checks_by_key),
            ),
            "CORR": self._build_dimension_result(
                dimension_id="CORR",
                label="Correctness",
                evidence_hints=self._collect_dimension_evidence(
                    checks_by_key,
                    keys=["risk_level_definition", "iso_14971_reference"],
                ),
                findings=self._collect_correctness_findings(fmea_precheck, rmf_precheck),
            ),
            "ADEQ": self._build_dimension_result(
                dimension_id="ADEQ",
                label="Adequacy",
                evidence_hints=self._collect_adequacy_evidence(fmea_rows, checks_by_key),
                findings=self._collect_adequacy_findings(fmea_rows, fmea_precheck),
            ),
            "TRAC": self._build_dimension_result(
                dimension_id="TRAC",
                label="Traceability",
                evidence_hints=self._collect_traceability_evidence(fmea_rows, checks_by_key, cross_doc_entities),
                findings=self._collect_traceability_findings(fmea_rows, fmea_precheck, rmf_precheck),
            ),
            "CONS": self._build_dimension_result(
                dimension_id="CONS",
                label="Consistency",
                evidence_hints=self._collect_consistency_evidence(checks_by_key, inventory_docs, term_map),
                findings=self._collect_consistency_findings(rmf_precheck, inventory_docs, term_map),
            ),
            "ACPT": self._build_dimension_result(
                dimension_id="ACPT",
                label="Acceptability",
                evidence_hints=self._collect_dimension_evidence(
                    checks_by_key,
                    keys=["residual_risk_evaluation", "risk_control_hierarchy"],
                ),
                findings=self._collect_acceptability_findings(run_manifest, fmea_precheck, rmf_precheck),
            ),
        }

        assessment = {
            "step_id": "rmf_dimension_review_agent",
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "document_ids": rmf_normalized.get("document_ids", []),
            "primary_review_object": rmf_normalized.get("primary_review_object"),
            "dimensions": dimensions,
            "global_manual_review_needed": True,
            "source_binding_enforced": run_manifest.get("workflow_name") is not None,
            "upstream_summary": {
                "fmea_row_count": len(fmea_rows),
                "fmea_finding_count": len(fmea_findings),
                "rmf_finding_count": len(rmf_findings),
                "rmf_precheck_status": rmf_precheck.get("structural_status"),
                "fmea_precheck_status": fmea_precheck.get("structural_status"),
            },
        }

        self._write_json(step_dir / "dimension_assessment.json", assessment)
        self._write_text(step_dir / "dimension_review_report.md", self._render_dimension_review_markdown(assessment))

    def _run_human_boundary(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("05_human_boundary")
        prompt_text = self._load_prompt_for_step(step)
        dimension_assessment = self._read_json(self._artifact_path("04_dimension_review", "dimension_assessment.json"))
        rmf_precheck = self._read_json(self._artifact_path("03_rmf_precheck", "rmf_precheck_report.json"))
        fmea_precheck = self._read_json(self._artifact_path("02_fmea_precheck", "fmea_precheck_report.json"))
        cross_doc_entities = self._read_json(self._artifact_path("01_parse", "cross_doc_entities.json"))

        dimensions = dimension_assessment.get("dimensions", {})
        upstream = dimension_assessment.get("upstream_summary", {})
        rmf_findings = rmf_precheck.get("findings", [])
        fmea_findings = fmea_precheck.get("findings", [])

        all_source_refs = self._collect_all_source_refs(dimension_assessment, rmf_precheck, fmea_precheck)

        review_items: list[dict[str, Any]] = []
        item_counter = 1

        comp_status = dimensions.get("COMP", {}).get("status", "insufficient_evidence")
        comp_findings = dimensions.get("COMP", {}).get("findings", [])
        if comp_findings or comp_status in ("issues_detected", "weakly_evidenced", "needs_human_review", "insufficient_evidence"):
            review_items.append(self._build_review_item(
                item_id=f"hrb_{item_counter:03d}",
                topic="risk_identification_adequacy",
                reviewer_focus="Completeness of risk identification: check whether all foreseeable hazards, hazardous situations, and harms have been identified in the RMF and FMEA.",
                why_not_auto_decidable="Machine cannot verify exhaustiveness of hazard identification; requires human expert judgment on whether the hazard set is complete.",
                priority=self._determine_priority(comp_findings, high_evidence_count=2),
                suggested_gate="conditional_pass",
                evidence_sources=self._extract_evidence_sources(comp_findings, all_source_refs),
            ))
            item_counter += 1

        fmea_row_count = upstream.get("fmea_row_count", 0)
        mapping_incomplete = fmea_precheck.get("mapping_incomplete_row_count", 0)
        if mapping_incomplete > 0 or fmea_findings:
            review_items.append(self._build_review_item(
                item_id=f"hrb_{item_counter:03d}",
                topic="probability_estimation_reasonableness",
                reviewer_focus=f"FMEA probability/occurrence scoring: verify that {mapping_incomplete} partially-mapped rows have been correctly assigned occurrence values; confirm scoring rationale is documented.",
                why_not_auto_decidable="Probability estimation requires domain knowledge and clinical context; automated mapping cannot validate reasonableness of assigned scores.",
                priority="high" if mapping_incomplete > 0 else "medium",
                suggested_gate="conditional_pass",
                evidence_sources=self._extract_evidence_sources(fmea_findings, all_source_refs),
            ))
            item_counter += 1

        corr_findings = dimensions.get("CORR", {}).get("findings", [])
        severity_findings = [f for f in corr_findings if "severity" in f.get("detail", "").lower() or "scoring" in f.get("detail", "").lower() or f.get("finding_type") == "method_scoring_uncertain"]
        if severity_findings or dimensions.get("CORR", {}).get("status") in ("issues_detected", "needs_human_review"):
            review_items.append(self._build_review_item(
                item_id=f"hrb_{item_counter:03d}",
                topic="severity_grading_appropriateness",
                reviewer_focus="Severity grading for each identified hazard: confirm that severity levels align with the defined severity classification scheme and clinical acceptability criteria.",
                why_not_auto_decidable="Severity grading is a clinical judgment that cannot be auto-verified without access to product-specific risk acceptance criteria and intended use context.",
                priority="high",
                suggested_gate="conditional_pass",
                evidence_sources=self._extract_evidence_sources(severity_findings, all_source_refs),
            ))
            item_counter += 1

        adeq_findings = dimensions.get("ADEQ", {}).get("findings", [])
        trac_findings = dimensions.get("TRAC", {}).get("findings", [])
        control_findings = [f for f in adeq_findings if "control" in f.get("detail", "").lower() or f.get("finding_type") in ("control_evidence_weak", "adequacy_not_demonstrated")]
        control_findings.extend([f for f in trac_findings if f.get("finding_type") == "trace_chain_incomplete"])
        if control_findings:
            review_items.append(self._build_review_item(
                item_id=f"hrb_{item_counter:03d}",
                topic="risk_control_adequacy",
                reviewer_focus="Risk control measures: verify that proposed controls are appropriate, implementation is verified, and effectiveness evidence is documented for each identified hazard.",
                why_not_auto_decidable="Adequacy of risk controls requires engineering and clinical judgment; automated checks cannot confirm control effectiveness.",
                priority="high",
                suggested_gate="conditional_pass",
                evidence_sources=self._extract_evidence_sources(control_findings, all_source_refs),
            ))
            item_counter += 1

        acpt_findings = dimensions.get("ACPT", {}).get("findings", [])
        residual_findings = [f for f in acpt_findings if "residual" in f.get("detail", "").lower() or f.get("finding_type") in ("human_decision_required", "acceptability_not_auto_releasable")]
        review_items.append(self._build_review_item(
            item_id=f"hrb_{item_counter:03d}",
            topic="residual_risk_acceptability",
            reviewer_focus="Residual risk acceptability: review the sponsor's justification for residual risk acceptability, verify alignment with stated risk acceptance criteria, and confirm no U-level risks remain unresolved.",
            why_not_auto_decidable="Residual risk acceptability is a deliberate human judgment that weighs residual risk against anticipated clinical benefit; it cannot be algorithmically determined.",
            priority="high",
            suggested_gate="rework_required" if any(f.get("finding_type") == "acceptability_not_auto_releasable" for f in acpt_findings) else "conditional_pass",
            evidence_sources=self._extract_evidence_sources(residual_findings, all_source_refs),
        ))
        item_counter += 1

        review_items.append(self._build_review_item(
            item_id=f"hrb_{item_counter:03d}",
            topic="overall_benefit_risk",
            reviewer_focus="Overall benefit-risk assessment: verify that the sponsor's overall benefit-risk determination is supported by evidence, consistent with intended use, and aligned with ISO 14971 residual risk evaluation requirements.",
            why_not_auto_decidable="Overall benefit-risk determination requires expert clinical and regulatory judgment; the runner cannot synthesize a benefit-risk conclusion.",
            priority="medium",
            suggested_gate="conditional_pass",
            evidence_sources=self._extract_evidence_sources(acpt_findings, all_source_refs),
        ))
        item_counter += 1

        cons_findings = dimensions.get("CONS", {}).get("findings", [])
        cross_doc_findings = [f for f in cons_findings if "term" in f.get("detail", "").lower() or f.get("finding_type") in ("term_alignment_weak", "cross_doc_correctness_weak", "ifu_missing")]
        cross_doc_findings.extend([f for f in corr_findings if f.get("finding_type") == "cross_doc_correctness_weak"])
        if cross_doc_findings or not cross_doc_entities.get("document_entities"):
            review_items.append(self._build_review_item(
                item_id=f"hrb_{item_counter:03d}",
                topic="cross_document_inconsistency_explanation",
                reviewer_focus="Cross-document consistency: review terminology alignment between RMF, FMEA, IFU, and other submitted documents; verify that inconsistent terminology has been explained or reconciled.",
                why_not_auto_decidable="Cross-document consistency requires semantic understanding and cannot be fully auto-verified; human review needed to confirm explanations are adequate.",
                priority="medium",
                suggested_gate="conditional_pass",
                evidence_sources=self._extract_evidence_sources(cross_doc_findings, all_source_refs),
            ))
            item_counter += 1

        review_items.append(self._build_review_item(
            item_id=f"hrb_{item_counter:03d}",
            topic="unknown_risk_evaluation",
            reviewer_focus="Unknown/unidentified risks: review whether the sponsor has adequately considered foreseeable but unidentified hazards, use errors, and production/post-production risks that could emerge.",
            why_not_auto_decidable="Unknown risks are, by definition, not present in the submitted documents; only human expert judgment can assess the adequacy of unknown-risk consideration.",
            priority="medium",
            suggested_gate="conditional_pass",
            evidence_sources=self._extract_evidence_sources(acpt_findings, all_source_refs),
        ))
        item_counter += 1

        gate = self._determine_provisional_gate(review_items, acpt_findings, upstream)

        human_review_queue = {
            "queue_id": f"hrq-{self.run_id}",
            "step_id": "rmf_human_boundary_agent",
            "prompt_contract_path": str((self.repo_root / step["prompt_contract"]).resolve()),
            "prompt_contract_loaded": bool(prompt_text.strip()),
            "recommended_gate": gate,
            "items": review_items,
            "upstream_summary": {
                "fmea_row_count": fmea_row_count,
                "fmea_finding_count": len(fmea_findings),
                "rmf_finding_count": len(rmf_findings),
                "total_review_items": len(review_items),
                "high_priority_items": sum(1 for item in review_items if item["priority"] == "high"),
                "dimension_assessment_status": dimension_assessment.get("overall_summary", {}).get("status", "unknown"),
            },
        }

        provisional_gate = {
            "step_id": "rmf_human_boundary_agent",
            "gate": gate,
            "basis": self._build_gate_basis(review_items, acpt_findings, upstream),
            "conditions_if_conditional": [
                item["reviewer_focus"] for item in review_items
                if item.get("suggested_gate_if_unresolved") == "conditional_pass"
            ] if gate == "conditional_pass" else [],
            "rework_triggers": [
                item["reviewer_focus"] for item in review_items
                if item.get("suggested_gate_if_unresolved") == "rework_required"
            ],
            "human_decision_required": True,
            "provisional_only": True,
            "caveat": "This recommendation is provisional and subject to human reviewer confirmation. It does not constitute a final regulatory decision.",
        }

        self._write_json(step_dir / "human_review_queue.json", human_review_queue)
        self._write_json(step_dir / "provisional_gate_recommendation.json", provisional_gate)

    def _build_review_item(
        self,
        *,
        item_id: str,
        topic: str,
        reviewer_focus: str,
        why_not_auto_decidable: str,
        priority: str,
        suggested_gate: str,
        evidence_sources: list[dict[str, Any]],
    ) -> dict[str, Any]:
        return {
            "item_id": item_id,
            "topic": topic,
            "reviewer_focus": reviewer_focus,
            "why_not_auto_decidable": why_not_auto_decidable,
            "priority": priority,
            "suggested_gate_if_unresolved": suggested_gate,
            "evidence_sources": evidence_sources if evidence_sources else [{"document_id": "upstream", "path": "", "section": "see_dimension_assessment", "quote": "Evidence captured in upstream dimension assessment findings"}],
        }

    def _determine_priority(self, findings: list[dict[str, Any]], *, high_evidence_count: int = 0) -> str:
        if any(f.get("severity") == "high" for f in findings):
            return "high"
        if len(findings) >= high_evidence_count:
            return "high"
        return "medium"

    def _extract_evidence_sources(self, findings: list[dict[str, Any]], fallback_refs: list[dict[str, Any]]) -> list[dict[str, Any]]:
        sources: list[dict[str, Any]] = []
        for finding in findings:
            for ref in finding.get("source_refs", []):
                if ref.get("document_id"):
                    sources.append(ref)
        return sources if sources else (fallback_refs[:2] if fallback_refs else [{"document_id": "upstream", "path": "", "section": "see_upstream_artifacts", "quote": "See upstream dimension and precheck artifacts"}])

    def _collect_all_source_refs(self, dimension_assessment: dict[str, Any], rmf_precheck: dict[str, Any], fmea_precheck: dict[str, Any]) -> list[dict[str, Any]]:
        refs: list[dict[str, Any]] = []
        for dim in dimension_assessment.get("dimensions", {}).values():
            for finding in dim.get("findings", []):
                refs.extend(finding.get("source_refs", []))
        for finding in rmf_precheck.get("findings", []):
            refs.extend(finding.get("source_refs", []))
        for finding in fmea_precheck.get("findings", []):
            refs.extend(finding.get("source_refs", []))
        return refs

    def _determine_provisional_gate(
        self,
        review_items: list[dict[str, Any]],
        acpt_findings: list[dict[str, Any]],
        upstream: dict[str, Any],
    ) -> str:
        high_priority_count = sum(1 for item in review_items if item["priority"] == "high")
        rework_required_items = [item for item in review_items if item.get("suggested_gate_if_unresolved") == "rework_required"]
        has_acceptability_issue = any(
            f.get("finding_type") in ("acceptability_not_auto_releasable", "human_decision_required")
            for f in acpt_findings
        )
        fmea_status = upstream.get("fmea_precheck_status", "unknown")
        rmf_status = upstream.get("rmf_precheck_status", "unknown")

        if has_acceptability_issue or len(rework_required_items) >= 2:
            return "rework_required"
        if high_priority_count >= 3 or fmea_status == "issues_detected" or rmf_status == "issues_detected":
            return "conditional_pass"
        return "conditional_pass"

    def _build_gate_basis(
        self,
        review_items: list[dict[str, Any]],
        acpt_findings: list[dict[str, Any]],
        upstream: dict[str, Any],
    ) -> str:
        high_priority = [item["item_id"] for item in review_items if item["priority"] == "high"]
        basis_parts = [
            f"Provisional gate based on {len(review_items)} human-boundary review items, "
            f"of which {len(high_priority)} are high-priority.",
        ]
        if upstream.get("fmea_finding_count", 0) > 0:
            basis_parts.append(f"FMEA precheck reported {upstream['fmea_finding_count']} findings.")
        if upstream.get("rmf_finding_count", 0) > 0:
            basis_parts.append(f"RMF precheck reported {upstream['rmf_finding_count']} findings.")
        if acpt_findings:
            basis_parts.append(f"ACPT dimension has {len(acpt_findings)} findings requiring human judgment.")
        basis_parts.append("Human reviewer must confirm, escalate, or resolve each queued item before any final gate decision.")
        return " ".join(basis_parts)

    def _run_report(self, step: dict[str, Any]) -> None:
        step_dir = self._artifact_dir("06_final")
        prompt_text = self._load_prompt_for_step(step)
        run_manifest = self._read_json(self._artifact_path("00_manifest", "run_manifest.json"))
        input_inventory = self._read_json(self._artifact_path("00_manifest", "input_inventory.json"))
        fmea_precheck = self._read_json(self._artifact_path("02_fmea_precheck", "fmea_precheck_report.json"))
        rmf_precheck = self._read_json(self._artifact_path("03_rmf_precheck", "rmf_precheck_report.json"))
        dimension_assessment = self._read_json(self._artifact_path("04_dimension_review", "dimension_assessment.json"))
        human_review_queue = self._read_json(self._artifact_path("05_human_boundary", "human_review_queue.json"))
        provisional_gate = self._read_json(self._artifact_path("05_human_boundary", "provisional_gate_recommendation.json"))

        capa_items = self._build_capa_list(fmea_precheck, rmf_precheck, dimension_assessment, human_review_queue)
        backflow_candidates = self._build_backflow_candidates(fmea_precheck, rmf_precheck, dimension_assessment, human_review_queue)

        recommended_gate = provisional_gate.get("gate", "rework_required")
        human_gate_status = "pending_human_confirmation"

        final_report_md = self._render_final_markdown(
            run_manifest=run_manifest,
            input_inventory=input_inventory,
            fmea_precheck=fmea_precheck,
            rmf_precheck=rmf_precheck,
            dimension_assessment=dimension_assessment,
            human_review_queue=human_review_queue,
            provisional_gate=provisional_gate,
            capa_items=capa_items,
            recommended_gate=recommended_gate,
            human_gate_status=human_gate_status,
        )

        final_report_json = {
            "step_id": "rmf_report_agent",
            "report_type": "rmf_review_final_report",
            "run_id": self.run_id,
            "thread_id": self.thread_id,
            "project_id": run_manifest.get("project_id"),
            "primary_review_object": run_manifest.get("primary_review_object"),
            "human_gate_required": run_manifest.get("human_gate_required", True),
            "recommended_gate": recommended_gate,
            "final_gate_status": human_gate_status,
            "gate_caveat": "MACHINE-GENERATED RECOMMENDATION ONLY. Final gate decision must be made by authorized human reviewer. This report does not constitute regulatory approval.",
            "executive_summary": self._build_executive_summary(fmea_precheck, rmf_precheck, dimension_assessment, human_review_queue, recommended_gate),
            "project_boundaries": {
                "project_id": run_manifest.get("project_id"),
                "project_name": run_manifest.get("project_id"),
                "institution_profile": run_manifest.get("institution_profile", {}),
                "primary_review_object": run_manifest.get("primary_review_object"),
                "review_scope": "single_project_serial_review",
                "assumptions": [
                    "All required documents were available or non-blocking per P0 configuration.",
                    "FMEA and Hazard Analysis treated as explicit review inputs.",
                    "Human gate is mandatory before any final compliance determination.",
                ],
            },
            "input_inventory_summary": {
                "total_documents": len(input_inventory.get("documents", [])),
                "present_count": sum(1 for d in input_inventory.get("documents", []) if d.get("status") == "present"),
                "missing_count": sum(1 for d in input_inventory.get("documents", []) if d.get("status") != "present"),
                "documents": input_inventory.get("documents", []),
            },
            "fmea_precheck_summary": {
                "row_count": fmea_precheck.get("row_count", 0),
                "duplicate_risk_ids": fmea_precheck.get("duplicate_risk_ids", []),
                "empty_row_count": fmea_precheck.get("empty_row_count", 0),
                "mapping_incomplete_row_count": fmea_precheck.get("mapping_incomplete_row_count", 0),
                "structural_status": fmea_precheck.get("structural_status", "unknown"),
                "findings_count": len(fmea_precheck.get("findings", [])),
                "findings": fmea_precheck.get("findings", []),
            },
            "rmf_precheck_summary": {
                "checks_evaluated": len(rmf_precheck.get("checks", [])),
                "evidence_hints_count": len(rmf_precheck.get("evidence_hints", [])),
                "structural_status": rmf_precheck.get("structural_status", "unknown"),
                "findings_count": len(rmf_precheck.get("findings", [])),
                "findings": rmf_precheck.get("findings", []),
                "checks": rmf_precheck.get("checks", []),
            },
            "dimension_review_summary": {
                "dimensions": {
                    dim_id: {
                        "status": dim.get("status"),
                        "label": dim.get("label"),
                        "findings_count": len(dim.get("findings", [])),
                        "findings": dim.get("findings", []),
                        "evidence_hints_count": len(dim.get("evidence_hints", [])),
                    }
                    for dim_id, dim in dimension_assessment.get("dimensions", {}).items()
                },
                "overall_status": dimension_assessment.get("overall_summary", {}).get("status", "unknown"),
                "global_manual_review_needed": dimension_assessment.get("global_manual_review_needed", True),
            },
            "human_boundary_summary": {
                "queue_id": human_review_queue.get("queue_id"),
                "recommended_gate": human_review_queue.get("recommended_gate"),
                "total_review_items": len(human_review_queue.get("items", [])),
                "high_priority_items": sum(1 for item in human_review_queue.get("items", []) if item.get("priority") == "high"),
                "review_items": human_review_queue.get("items", []),
            },
            "provisional_gate_recommendation": provisional_gate,
            "capa_action_list": capa_items,
            "backflow_candidates": backflow_candidates,
            "source_artifacts": {
                "run_manifest": str(self._artifact_path("00_manifest", "run_manifest.json")),
                "input_inventory": str(self._artifact_path("00_manifest", "input_inventory.json")),
                "fmea_precheck_report": str(self._artifact_path("02_fmea_precheck", "fmea_precheck_report.json")),
                "rmf_precheck_report": str(self._artifact_path("03_rmf_precheck", "rmf_precheck_report.json")),
                "dimension_assessment": str(self._artifact_path("04_dimension_review", "dimension_assessment.json")),
                "dimension_review_report": str(self._artifact_path("04_dimension_review", "dimension_review_report.md")),
                "human_review_queue": str(self._artifact_path("05_human_boundary", "human_review_queue.json")),
                "provisional_gate_recommendation": str(self._artifact_path("05_human_boundary", "provisional_gate_recommendation.json")),
            },
        }

        self._write_text(step_dir / "final_report.md", final_report_md)
        self._write_json(step_dir / "final_report.json", final_report_json)
        self._write_json(step_dir / "capa_action_list.json", capa_items)
        self._write_json(step_dir / "backflow_candidates.json", backflow_candidates)

    def _build_executive_summary(
        self,
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
        dimension_assessment: dict[str, Any],
        human_review_queue: dict[str, Any],
        recommended_gate: str,
    ) -> dict[str, Any]:
        return {
            "review_conclusion": f"Machine-generated recommendation: {recommended_gate}",
            "fmea_status": fmea_precheck.get("structural_status", "unknown"),
            "rmf_status": rmf_precheck.get("structural_status", "unknown"),
            "review_item_count": len(human_review_queue.get("items", [])),
            "high_priority_count": sum(1 for item in human_review_queue.get("items", []) if item.get("priority") == "high"),
            "human_gate_required": True,
            "machine_finding": "Machine-generated recommendation only. Await human gate confirmation.",
        }

    def _build_capa_list(
        self,
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
        dimension_assessment: dict[str, Any],
        human_review_queue: dict[str, Any],
    ) -> list[dict[str, Any]]:
        capa_items: list[dict[str, Any]] = []
        capa_counter = 1

        for finding in fmea_precheck.get("findings", []):
            finding_type = finding.get("finding_type", "unknown")
            if finding_type == "duplicate_risk_id":
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "fmea_precheck",
                    "source_step": "fmea_precheck_agent",
                    "finding_type": finding_type,
                    "description": f"Duplicate risk ID '{finding.get('risk_id')}' found in FMEA. Deduplicate or clarify risk ID assignment.",
                    "blocking": True,
                    "priority": "high",
                    "suggested_action": "Deduplicate risk IDs in FMEA; ensure each risk has a unique identifier.",
                    "evidence_refs": [finding.get("source_ref", {})],
                })
                capa_counter += 1
            elif finding_type == "mapping_incomplete":
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "fmea_precheck",
                    "source_step": "fmea_precheck_agent",
                    "finding_type": finding_type,
                    "description": f"Row index {finding.get('row_index')}: semantic mapping incomplete for fields {finding.get('missing_fields', [])}.",
                    "blocking": True,
                    "priority": "high",
                    "suggested_action": "Complete semantic mapping for affected rows; verify all fields are correctly interpreted.",
                    "evidence_refs": [finding.get("source_ref", {})],
                })
                capa_counter += 1
            elif finding_type == "missing_required_fields":
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "fmea_precheck",
                    "source_step": "fmea_precheck_agent",
                    "finding_type": finding_type,
                    "description": f"Row index {finding.get('row_index')} (risk_id={finding.get('risk_id') or 'N/A'}): missing required fields: {finding.get('missing_fields', [])}.",
                    "blocking": True,
                    "priority": "high",
                    "suggested_action": "Populate all required FMEA fields for the affected row.",
                    "evidence_refs": [finding.get("source_ref", {})],
                })
                capa_counter += 1

        for finding in rmf_precheck.get("findings", []):
            finding_type = finding.get("finding_type", "unknown")
            if finding_type in ("weak_evidence", "coverage_gap"):
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "rmf_precheck",
                    "source_step": "rmf_precheck_agent",
                    "finding_type": finding_type,
                    "description": f"RMF section check '{finding.get('check_key', 'unknown')}': {finding.get('detail', 'weak evidence for required section')}.",
                    "blocking": finding.get("severity") == "high",
                    "priority": "medium" if finding.get("severity") != "high" else "high",
                    "suggested_action": "Strengthen evidence for the indicated RMF section; add explicit section content or cross-reference.",
                    "evidence_refs": finding.get("source_refs", []),
                })
                capa_counter += 1
            elif finding_type == "manual_followup_required":
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "rmf_precheck",
                    "source_step": "rmf_precheck_agent",
                    "finding_type": finding_type,
                    "description": finding.get("label", "Manual follow-up required for RMF-FMEA traceability."),
                    "blocking": False,
                    "priority": "medium",
                    "suggested_action": "Verify FMEA-to-RMF traceability manually; document cross-reference in RMF.",
                    "evidence_refs": finding.get("source_refs", []),
                })
                capa_counter += 1

        for dim_id, dim in dimension_assessment.get("dimensions", {}).items():
            for finding in dim.get("findings", []):
                finding_type = finding.get("finding_type", "unknown")
                capa_items.append({
                    "capa_id": f"capa_{capa_counter:03d}",
                    "source": "dimension_review",
                    "source_step": "rmf_dimension_review_agent",
                    "dimension": dim_id,
                    "finding_type": finding_type,
                    "description": f"[{dim_id}] {finding.get('detail', 'Dimension finding requires remediation.')}",
                    "blocking": dim.get("status") in ("issues_detected",) and finding_type in ("human_decision_required", "acceptability_not_auto_releasable"),
                    "priority": "high" if dim.get("status") == "issues_detected" else "medium",
                    "suggested_action": self._suggest_action_for_finding(finding_type, dim_id),
                    "evidence_refs": finding.get("source_refs", []),
                })
                capa_counter += 1

        for item in human_review_queue.get("items", []):
            topic = item.get("topic", "unknown")
            priority = item.get("priority", "medium")
            capa_items.append({
                "capa_id": f"capa_{capa_counter:03d}",
                "source": "human_boundary",
                "source_step": "rmf_human_boundary_agent",
                "review_item_id": item.get("item_id"),
                "topic": topic,
                "description": f"[Human Review Required] {item.get('reviewer_focus', 'Human expert judgment needed.')}",
                "blocking": priority == "high" or item.get("suggested_gate_if_unresolved") == "rework_required",
                "priority": priority,
                "suggested_action": item.get("reviewer_focus", "Human reviewer must evaluate and confirm."),
                "evidence_refs": item.get("evidence_sources", []),
            })
            capa_counter += 1

        return capa_items

    def _build_backflow_candidates(
        self,
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
        dimension_assessment: dict[str, Any],
        human_review_queue: dict[str, Any],
    ) -> list[dict[str, Any]]:
        candidates: list[dict[str, Any]] = []
        bc_counter = 1

        if fmea_precheck.get("mapping_incomplete_row_count", 0) > 0:
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "fmea_mapping_gap",
                "category": "mapping_gaps",
                "description": f"{fmea_precheck['mapping_incomplete_row_count']} FMEA rows have incomplete semantic mapping; fields not auto-interpretable.",
                "root_cause": "Source FMEA table format does not match expected column aliases; raw_cells retained.",
                "recommendation": "Standardize FMEA column headers against schema; add header normalization step in runner pre-processing.",
                "reusable": True,
                "affects_review_object": "FMEA",
            })
            bc_counter += 1

        duplicate_count = len(fmea_precheck.get("duplicate_risk_ids", []))
        if duplicate_count > 0:
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "duplicate_risk_id",
                "category": "failure_patterns",
                "description": f"{duplicate_count} duplicate risk IDs found across FMEA rows.",
                "root_cause": "Risk ID generation is not enforced to be unique in the source document.",
                "recommendation": "Introduce a risk ID uniqueness check at FMEA intake; enforce uniqueness constraint.",
                "reusable": True,
                "affects_review_object": "FMEA",
            })
            bc_counter += 1

        if fmea_precheck.get("orphan_row_count", 0) > 0:
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "orphan_fmea_rows",
                "category": "failure_patterns",
                "description": f"{fmea_precheck['orphan_row_count']} FMEA rows have risk IDs but no assigned controls.",
                "root_cause": "FMEA row has a risk ID but no control measure recorded.",
                "recommendation": "Require control measure field to be populated for every risk ID row; add validation at FMEA precheck.",
                "reusable": True,
                "affects_review_object": "FMEA",
            })
            bc_counter += 1

        weak_evidence_findings = [f for f in rmf_precheck.get("findings", []) if f.get("finding_type") == "weak_evidence"]
        if weak_evidence_findings:
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "weak_evidence_across_rmf_sections",
                "category": "reviewer_judgment_hotspots",
                "description": f"{len(weak_evidence_findings)} RMF sections have weak evidence; reviewer must manually verify presence.",
                "root_cause": "Section detection relies on keyword matching which may miss non-standard section titles.",
                "recommendation": "Improve section detection by adding fuzzy matching or alternative keywords; or require explicit section tagging in source documents.",
                "reusable": True,
                "affects_review_object": "RMF",
            })
            bc_counter += 1

        term_consistency = rmf_precheck.get("term_consistency", {})
        if term_consistency.get("status") != "present":
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "terminology_inconsistency",
                "category": "reviewer_judgment_hotspots",
                "description": "RMF vs IFU terminology alignment is weak; cross-document consistency review needed.",
                "root_cause": "IFU and RMF may use different terminology for the same concepts; automated cross-reference is incomplete.",
                "recommendation": "Add explicit terminology mapping between IFU and RMF; require cross-reference table in RMF annex.",
                "reusable": True,
                "affects_review_object": "RMF,IFU",
            })
            bc_counter += 1

        high_priority_human_items = [item for item in human_review_queue.get("items", []) if item.get("priority") == "high"]
        if high_priority_human_items:
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "human_boundary_patterns",
                "category": "reviewer_judgment_hotspots",
                "description": f"{len(high_priority_human_items)} high-priority human review items identified; these represent systematic judgment requirements.",
                "root_cause": "Systematic: probability estimation, severity grading, residual risk acceptability, and control adequacy require domain expertise beyond rule-based checking.",
                "recommendation": "For future reviews: consider adding structured guidance documents for probability/severity scoring criteria to reduce human judgment variability.",
                "reusable": True,
                "affects_review_object": "FMEA,RMF",
            })
            bc_counter += 1

        if dimension_assessment.get("global_manual_review_needed"):
            candidates.append({
                "bc_id": f"bc_{bc_counter:03d}",
                "pattern_type": "global_manual_review_needed",
                "category": "reviewer_judgment_hotspots",
                "description": "Global manual review flag is set; at least one dimension requires human expert assessment.",
                "root_cause": "At least one dimension has issues_detected status that cannot be auto-resolved.",
                "recommendation": "Establish a reviewer checklist that maps dimension statuses to required reviewer expertise.",
                "reusable": True,
                "affects_review_object": "RMF,FMEA",
            })
            bc_counter += 1

        return candidates

    def _suggest_action_for_finding(self, finding_type: str, dim_id: str) -> str:
        action_map = {
            "human_decision_required": "Human reviewer must make explicit decision; no automation possible.",
            "acceptability_not_auto_releasable": "Auto-release not permitted; human reviewer must confirm acceptability.",
            "adequacy_not_demonstrated": "Provide additional evidence demonstrating adequacy; supplement FMEA or RMF section.",
            "traceability_gap": "Strengthen traceability between RMF and FMEA; add explicit cross-references.",
            "trace_chain_incomplete": "Complete the risk-control-verification chain for affected rows.",
            "term_alignment_weak": "Align terminology between RMF and IFU; document mapping or reconcile naming.",
            "control_evidence_weak": "Provide stronger evidence of control implementation and verification.",
            "method_scoring_uncertain": "Clarify FMEA scoring methodology; document rationale for probability and severity assignments.",
            "cross_doc_correctness_weak": "Verify cross-document consistency for the affected terms or sections.",
            "coverage_gap": "Add or strengthen evidence for the missing coverage area.",
            "weak_terminology_alignment": "Strengthen RMF terminology baseline; cross-reference with IFU.",
        }
        return action_map.get(
            finding_type,
            f"Review [{dim_id}] finding of type '{finding_type}'; apply domain expertise to resolve.",
        )

    def _render_final_markdown(
        self,
        *,
        run_manifest: dict[str, Any],
        input_inventory: dict[str, Any],
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
        dimension_assessment: dict[str, Any],
        human_review_queue: dict[str, Any],
        provisional_gate: dict[str, Any],
        capa_items: list[dict[str, Any]],
        recommended_gate: str,
        human_gate_status: str,
    ) -> str:
        lines: list[str] = []
        lines.append("# RMF Review Final Report")
        lines.append("")
        lines.append(f"**⚠️  MACHINE-GENERATED REPORT — FINAL GATE DECISION PENDING HUMAN CONFIRMATION**")
        lines.append("")
        lines.append(f"- Run ID: `{self.run_id}`")
        lines.append(f"- Project ID: `{run_manifest.get('project_id', 'unknown')}`")
        lines.append(f"- Primary Review Object: `{run_manifest.get('primary_review_object', 'RMF')}`")
        lines.append(f"- Human Gate Required: `True`")
        lines.append(f"- Machine Recommendation: **{recommended_gate.upper()}**")
        lines.append(f"- Final Gate Status: `{human_gate_status}`")
        lines.append("")

        lines.append("---")
        lines.append("## 1. Executive Summary")
        lines.append("")
        lines.append(f"This report covers the RMF review for project `{run_manifest.get('project_id')}`.")
        lines.append(f"FMEA precheck identified **{fmea_precheck.get('row_count', 0)}** rows with **{len(fmea_precheck.get('findings', []))}** findings.")
        lines.append(f"RMF precheck identified **{len(rmf_precheck.get('findings', []))}** findings across **{len(rmf_precheck.get('checks', []))}** section checks.")
        lines.append(f"The six-dimension review flagged **{sum(1 for d in dimension_assessment.get('dimensions', {}).values() if d.get('status') != 'supported')}** dimensions with issues.")
        lines.append(f"Human boundary review identified **{len(human_review_queue.get('items', []))}** review items, of which **{sum(1 for i in human_review_queue.get('items', []) if i.get('priority') == 'high')}** are high-priority.")
        lines.append(f"")
        lines.append(f"**Machine-Recommended Gate: `{recommended_gate}`** — *awaiting human confirmation*.")
        lines.append("")

        lines.append("---")
        lines.append("## 2. Project Boundaries and Assumptions")
        lines.append("")
        lines.append(f"- Institution Profile: `{run_manifest.get('institution_profile', {}).get('assessment_body', 'BSI')}`")
        lines.append(f"- Primary Review Object: `{run_manifest.get('primary_review_object', 'RMF')}`")
        lines.append(f"- Review Mode: `single_project_serial_review`")
        lines.append(f"- Human Gate Required: `True` (non-negotiable)")
        lines.append("")
        lines.append("**Assumptions:**")
        lines.append("1. All required documents were available or downgraded to non-blocking per P0 configuration.")
        lines.append("2. FMEA and Hazard Analysis treated as explicit review inputs.")
        lines.append("3. No source document was modified during this review cycle.")
        lines.append("")

        lines.append("---")
        lines.append("## 3. Input Inventory Summary")
        lines.append("")
        docs = input_inventory.get("documents", [])
        lines.append(f"Total documents declared: **{len(docs)}**")
        present = [d for d in docs if d.get("status") == "present"]
        missing = [d for d in docs if d.get("status") != "present"]
        lines.append(f"- Present: {len(present)}")
        lines.append(f"- Missing or degraded: {len(missing)}")
        lines.append("")
        for doc in docs:
            status_icon = "✅" if doc.get("status") == "present" else "⚠️"
            lines.append(f"{status_icon} `{doc.get('doc_type')}` — {doc.get('label')} — `{doc.get('status')}`")
        lines.append("")

        lines.append("---")
        lines.append("## 4. FMEA Precheck Summary")
        lines.append("")
        lines.append(f"- Row Count: `{fmea_precheck.get('row_count', 0)}`")
        lines.append(f"- Duplicate Risk IDs: `{len(fmea_precheck.get('duplicate_risk_ids', []))}`")
        lines.append(f"- Empty Rows: `{fmea_precheck.get('empty_row_count', 0)}`")
        lines.append(f"- Mapping Incomplete Rows: `{fmea_precheck.get('mapping_incomplete_row_count', 0)}`")
        lines.append(f"- Structural Status: `{fmea_precheck.get('structural_status', 'unknown')}`")
        lines.append(f"- Findings Count: `{len(fmea_precheck.get('findings', []))}`")
        lines.append("")
        if fmea_precheck.get("findings"):
            lines.append("### FMEA Findings")
            for f in fmea_precheck.get("findings", []):
                lines.append(f"- `{f.get('finding_type')}` | row={f.get('row_index', 'N/A')} | risk_id={f.get('risk_id') or 'N/A'} | {f.get('detail', '')}")
            lines.append("")
        else:
            lines.append("No FMEA findings.")
            lines.append("")

        lines.append("---")
        lines.append("## 5. RMF Precheck Summary")
        lines.append("")
        lines.append(f"- Checks Evaluated: `{len(rmf_precheck.get('checks', []))}`")
        lines.append(f"- Evidence Hints: `{len(rmf_precheck.get('evidence_hints', []))}`")
        lines.append(f"- Structural Status: `{rmf_precheck.get('structural_status', 'unknown')}`")
        lines.append(f"- Findings Count: `{len(rmf_precheck.get('findings', []))}`")
        lines.append("")
        lines.append("### RMF Section Checks")
        for check in rmf_precheck.get("checks", []):
            lines.append(f"- `{check.get('check_key')}` | state=`{check.get('state')}`")
        lines.append("")
        if rmf_precheck.get("findings"):
            lines.append("### RMF Findings")
            for f in rmf_precheck.get("findings", []):
                lines.append(f"- `{f.get('finding_type')}` | `{f.get('check_key', 'n/a')}` | {f.get('detail', '')}")
            lines.append("")
        else:
            lines.append("No RMF findings.")
            lines.append("")

        lines.append("---")
        lines.append("## 6. Six-Dimension Review Summary")
        lines.append("")
        dim_map = {
            "COMP": "Completeness",
            "CORR": "Correctness",
            "ADEQ": "Adequacy",
            "TRAC": "Traceability",
            "CONS": "Consistency",
            "ACPT": "Acceptability",
        }
        for dim_id, dim_label in dim_map.items():
            dim_data = dimension_assessment.get("dimensions", {}).get(dim_id, {})
            status = dim_data.get("status", "unknown")
            status_icon = {"supported": "✅", "issues_detected": "⚠️", "weakly_evidenced": "⚠️", "needs_human_review": "🔶", "insufficient_evidence": "❌"}.get(status, "❓")
            lines.append(f"{status_icon} **{dim_id}** ({dim_label}): status=`{status}`, findings={len(dim_data.get('findings', []))}")
        lines.append("")
        lines.append(f"Global Manual Review Needed: `{dimension_assessment.get('global_manual_review_needed', True)}`")
        lines.append("")

        lines.append("---")
        lines.append("## 7. Human Boundary Summary")
        lines.append("")
        items = human_review_queue.get("items", [])
        high_items = [i for i in items if i.get("priority") == "high"]
        lines.append(f"- Queue ID: `{human_review_queue.get('queue_id')}`")
        lines.append(f"- Recommended Gate: `{human_review_queue.get('recommended_gate')}`")
        lines.append(f"- Total Review Items: `{len(items)}`")
        lines.append(f"- High Priority Items: `{len(high_items)}`")
        lines.append("")
        lines.append("### Human Review Items")
        for item in items:
            lines.append(f"- `[{'🔴' if item.get('priority') == 'high' else '🟡'}]` **`{item.get('item_id')}`** topic=`{item.get('topic')}`")
            lines.append(f"  - Focus: {item.get('reviewer_focus', '')[:120]}")
            lines.append(f"  - Why not auto: {item.get('why_not_auto_decidable', '')[:100]}")
            lines.append(f"  - Suggested gate if unresolved: `{item.get('suggested_gate_if_unresolved')}`")
        lines.append("")

        lines.append("---")
        lines.append("## 8. Provisional Gate Recommendation")
        lines.append("")
        lines.append(f"**Machine-Generated Recommendation: `{provisional_gate.get('gate')}`**")
        lines.append("")
        lines.append(f"Basis: {provisional_gate.get('basis', 'See provisional_gate_recommendation artifact.')}")
        lines.append("")
        lines.append("⚠️  **This is a machine-generated provisional recommendation only.**")
        lines.append("The final gate decision must be made by an authorized human reviewer.")
        lines.append("Do not treat this document as regulatory approval or final compliance determination.")
        lines.append("")

        if provisional_gate.get("conditions_if_conditional"):
            lines.append("**Conditions for conditional pass:**")
            for cond in provisional_gate.get("conditions_if_conditional", []):
                lines.append(f"- {cond}")
            lines.append("")

        if provisional_gate.get("rework_triggers"):
            lines.append("**Rework triggers:**")
            for trigger in provisional_gate.get("rework_triggers", []):
                lines.append(f"- {trigger}")
            lines.append("")

        lines.append("---")
        lines.append("## 9. CAPA Action List")
        lines.append("")
        lines.append(f"Total CAPA items: **{len(capa_items)}**")
        blocking = [c for c in capa_items if c.get("blocking")]
        non_blocking = [c for c in capa_items if not c.get("blocking")]
        lines.append(f"- Blocking: **{len(blocking)}**")
        lines.append(f"- Non-blocking: **{len(non_blocking)}**")
        lines.append("")
        lines.append("### Blocking CAPA Items")
        if blocking:
            for capa in blocking:
                lines.append(f"- **`{capa.get('capa_id')}`** [{capa.get('source')}] {capa.get('description', '')[:150]}")
                lines.append(f"  - Priority: `{capa.get('priority')}`")
                lines.append(f"  - Suggested action: {capa.get('suggested_action', '')}")
        else:
            lines.append("None")
        lines.append("")
        lines.append("### Non-Blocking CAPA Items")
        if non_blocking:
            for capa in non_blocking:
                lines.append(f"- **`{capa.get('capa_id')}`** [{capa.get('source')}] {capa.get('description', '')[:150]}")
        else:
            lines.append("None")
        lines.append("")

        lines.append("---")
        lines.append("## 10. Backflow Candidates")
        lines.append("")
        lines.append("Backflow candidates are reusable patterns identified during this review that can inform future review improvements.")
        lines.append("")
        for bc in self._build_backflow_candidates(fmea_precheck, rmf_precheck, dimension_assessment, human_review_queue):
            lines.append(f"- **`{bc.get('bc_id')}`** pattern=`{bc.get('pattern_type')}` category=`{bc.get('category')}`")
            lines.append(f"  - {bc.get('description', '')[:150]}")
            lines.append(f"  - Recommendation: {bc.get('recommendation', '')[:120]}")
        lines.append("")

        lines.append("---")
        lines.append("## Source Artifacts")
        lines.append("")
        lines.append(f"- Run manifest: `00_manifest/run_manifest.json`")
        lines.append(f"- Input inventory: `00_manifest/input_inventory.json`")
        lines.append(f"- FMEA precheck: `02_fmea_precheck/fmea_precheck_report.json`")
        lines.append(f"- RMF precheck: `03_rmf_precheck/rmf_precheck_report.json`")
        lines.append(f"- Dimension assessment: `04_dimension_review/dimension_assessment.json`")
        lines.append(f"- Human review queue: `05_human_boundary/human_review_queue.json`")
        lines.append(f"- Provisional gate: `05_human_boundary/provisional_gate_recommendation.json`")
        lines.append("")
        lines.append(f"*Report generated by RMFReviewRunner v1.1 — run_id: {self.run_id} — thread_id: {self.thread_id}*")
        lines.append("")

        return "\n".join(lines)

    def _build_dry_run_plan(self) -> dict[str, Any]:
        return {
            "workflow_name": self.workflow_name,
            "run_id": self.run_id,
            "thread_id": self.thread_id,
            "workflow_path": str(self.workflow_path),
            "project_profile_path": str(self.project_profile_path),
            "input_root": str(self._resolve_input_root()),
            "artifact_root_virtual": self.artifact_root_virtual,
            "artifact_root_actual": str(self.artifact_root_actual),
            "supported_steps": list(_SUPPORTED_STEP_IDS),
            "ordered_steps": [step.get("step_id") for step in self.workflow.get("ordered_steps", [])],
        }

    def _build_fmea_normalized(self, fmea_docs: list[dict[str, Any]]) -> dict[str, Any]:
        rows: list[dict[str, Any]] = []
        normalization_notes = [
            "Generated by minimal RMF runner glue.",
            "Rows may be partially normalized when source tables are heterogeneous.",
        ]
        for item in fmea_docs:
            path = Path(item["resolved_path"])
            extracted_rows = self._extract_tabular_rows(path)
            if path.suffix.lower() == ".docx":
                normalization_notes.append(f"DOCX table extraction enabled for {item['document_id']}.")
            mapped_rows = self._map_fmea_rows(extracted_rows, item)
            rows.extend(mapped_rows)
            if extracted_rows and not any(row.get("risk_id") for row in mapped_rows):
                normalization_notes.append(f"Field mapping incomplete for {item['document_id']}; raw_cells retained for manual review.")

        return {
            "document_id": "fmea_bundle",
            "document_type": "Combined_FMEA_Hazard_Analysis",
            "normalization_notes": normalization_notes,
            "rows": rows,
        }

    def _build_rmf_normalized(self, rmf_docs: list[dict[str, Any]]) -> dict[str, Any]:
        all_text_lines: list[dict[str, Any]] = []
        for item in rmf_docs:
            all_text_lines.extend(self._extract_document_text_hints(Path(item["resolved_path"]), item))

        section_specs = [
            ("risk_management_plan", ["风险管理计划", "风险管理报告", "risk management plan", "risk management report"]),
            ("risk_analysis_matrix", ["风险分析", "风险评价准则", "probability", "severity", "risk analysis"]),
            ("risk_level_definition", ["严重度", "发生概率", "发生频度", "severity", "probability"]),
            ("traceability_matrix", ["可追溯", "traceability"]),
            ("residual_risk_evaluation", ["剩余风险", "综合剩余风险", "residual risk"]),
            ("production_post_production_information", ["生产和生产后", "生产后信息", "post-production"]),
            ("iso_14971_reference", ["iso 14971", "iso14971", "yy/t 0316"]),
            ("risk_control_hierarchy", ["固有安全设计", "防护措施", "安全信息", "三步法"]),
        ]
        section_hints = []
        for section_key, keywords in section_specs:
            match = self._search_text_hints(all_text_lines, keywords)
            section_hints.append(
                {
                    "section_key": section_key,
                    "status": "hinted" if match else "candidate_only",
                    "evidence": match["evidence"] if match else [],
                    "source_refs": match["source_refs"] if match else [item["source_ref"] for item in rmf_docs],
                }
            )

        return {
            "document_ids": [item["document_id"] for item in rmf_docs],
            "document_paths": [item["resolved_path"] for item in rmf_docs],
            "primary_review_object": "RMF",
            "sections": self._infer_rmf_sections(rmf_docs),
            "section_hints": section_hints,
            "text_lines": all_text_lines,
            "source_refs": [item["source_ref"] for item in rmf_docs],
        }

    def _extract_tabular_rows(self, path: Path) -> list[dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle, delimiter=delimiter)
                return [dict(row) for row in reader]

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except Exception:
                logger.warning("openpyxl unavailable, skipping xlsx parsing for %s", path)
                return []

            workbook = load_workbook(path, read_only=True, data_only=True)
            all_rows: list[dict[str, Any]] = []
            for sheet in workbook.worksheets:
                raw_rows = list(sheet.iter_rows(values_only=True))
                header_index = self._find_header_row(raw_rows)
                if header_index is None:
                    continue
                headers = [self._clean_cell(value) for value in raw_rows[header_index]]
                for row_idx, row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
                    values = [self._clean_cell(value) for value in row]
                    if not any(values):
                        continue
                    record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
                    record["__sheet_name"] = sheet.title
                    record["__sheet_row"] = row_idx
                    all_rows.append(record)
            return all_rows

        if suffix == ".docx":
            return self._extract_docx_tabular_rows(path)

        return []

    def _extract_docx_tabular_rows(self, path: Path) -> list[dict[str, Any]]:
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        candidate_tables: list[tuple[int, list[dict[str, Any]]]] = []
        fallback_tables: list[list[dict[str, Any]]] = []
        try:
            with ZipFile(path) as archive:
                document_xml = archive.read("word/document.xml")
        except Exception:
            logger.warning("Unable to read docx file for table extraction: %s", path, exc_info=True)
            return []

        root = ET.fromstring(document_xml)
        for table_index, table in enumerate(root.findall(".//w:tbl", namespace), start=1):
            table_rows: list[list[str]] = []
            for row in table.findall("./w:tr", namespace):
                cells: list[str] = []
                for cell in row.findall("./w:tc", namespace):
                    text_parts = [node.text for node in cell.findall(".//w:t", namespace) if node.text]
                    cell_text = " ".join(part.strip() for part in text_parts if part and part.strip()).strip()
                    cells.append(cell_text)
                if any(cells):
                    table_rows.append(cells)

            if not table_rows:
                continue

            header_index = self._find_header_row([tuple(row) for row in table_rows])
            table_records: list[dict[str, Any]] = []
            if header_index is None:
                headers = [f"col_{i + 1}" for i in range(max(len(row) for row in table_rows))]
                data_rows = table_rows
            else:
                headers = [self._clean_cell(value) or f"col_{idx + 1}" for idx, value in enumerate(table_rows[header_index])]
                data_rows = table_rows[header_index + 1 :]

            for row_idx, values in enumerate(data_rows, start=1 if header_index is None else header_index + 2):
                padded_values = list(values) + [""] * max(0, len(headers) - len(values))
                if not any(cell.strip() for cell in padded_values):
                    continue
                record = {headers[i]: padded_values[i] if i < len(padded_values) else "" for i in range(len(headers))}
                record["__table_name"] = f"table_{table_index}"
                record["__sheet_name"] = f"table_{table_index}"
                record["__sheet_row"] = row_idx
                record["__raw_cells"] = list(values)
                table_records.append(record)

            if not table_records:
                continue
            if header_index is None:
                fallback_tables.append(table_records)
            else:
                candidate_tables.append((self._score_header_cells(headers), table_records))

        if candidate_tables:
            best_score = max(score for score, _ in candidate_tables)
            selected_tables = [table for score, table in candidate_tables if score == best_score]
            return [row for table in selected_tables for row in table]
        return [row for table in fallback_tables for row in table]

    def _extract_document_text_hints(self, path: Path, item: dict[str, Any]) -> list[dict[str, Any]]:
        suffix = path.suffix.lower()
        if suffix == ".docx":
            return self._extract_docx_text_hints(path, item)
        if suffix in {".txt", ".md"}:
            lines = [line.strip() for line in path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]
            return [
                {
                    "text": line,
                    "source_ref": {"document_id": item["document_id"], "path": item["declared_path"], "line_id": str(index)},
                }
                for index, line in enumerate(lines[:400], start=1)
            ]
        return [
            {
                "text": path.name,
                "source_ref": {"document_id": item["document_id"], "path": item["declared_path"], "line_id": "filename"},
            }
        ]

    def _extract_docx_text_hints(self, path: Path, item: dict[str, Any]) -> list[dict[str, Any]]:
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        try:
            with ZipFile(path) as archive:
                document_xml = archive.read("word/document.xml")
        except Exception:
            logger.warning("Unable to read docx text for hint extraction: %s", path, exc_info=True)
            return []

        root = ET.fromstring(document_xml)
        text_lines: list[dict[str, Any]] = []
        for index, para in enumerate(root.findall(".//w:p", namespace), start=1):
            text = "".join(node.text for node in para.findall(".//w:t", namespace) if node.text).strip()
            if not text:
                continue
            text_lines.append(
                {
                    "text": text,
                    "source_ref": {"document_id": item["document_id"], "path": item["declared_path"], "line_id": str(index)},
                }
            )
        return text_lines[:500]

    def _map_fmea_rows(self, rows: list[dict[str, Any]], item: dict[str, Any]) -> list[dict[str, Any]]:
        mapped: list[dict[str, Any]] = []
        for index, row in enumerate(rows, start=1):
            norm = {self._normalize_key(key): value for key, value in row.items()}
            raw_cells = row.get("__raw_cells") or [value for key, value in row.items() if not str(key).startswith("__")]
            risk_id = self._find_value(norm, "risk_id")
            failure_mode = self._find_value(norm, "failure_mode")
            effect = self._find_value(norm, "effect")
            cause = self._find_value(norm, "cause")
            current_control = self._find_value(norm, "current_control")
            probability_value = self._find_value(norm, "probability")
            occurrence_value = self._find_value(norm, "occurrence")
            severity_value = self._find_value(norm, "severity")
            detection_value = self._find_value(norm, "detection")
            risk_level_value = self._find_value(norm, "risk_level")
            rpn_value = self._find_value(norm, "rpn")
            mapped_fields = [
                risk_id,
                failure_mode,
                effect,
                cause,
                current_control,
                probability_value,
                occurrence_value,
                severity_value,
                detection_value,
                risk_level_value,
                rpn_value,
            ]
            mapped_row = {
                "risk_id": risk_id,
                "hazard": self._find_value(norm, "hazard"),
                "hazardous_situation": self._find_value(norm, "hazardous_situation"),
                "harm": self._find_value(norm, "harm"),
                "failure_mode": failure_mode,
                "effect": effect,
                "cause": cause,
                "sequence_of_events": self._split_list_value(self._find_value(norm, "sequence_of_events")),
                "probability": self._metric(probability_value, item),
                "occurrence": self._metric(occurrence_value, item),
                "severity": self._metric(severity_value, item),
                "detection": self._metric(detection_value, item),
                "risk_level": self._metric(risk_level_value, item),
                "rpn": self._metric(rpn_value, item),
                "current_control": current_control,
                "controls": self._controls(self._find_value(norm, "controls") or current_control, item, index),
                "acceptance_conclusion": {
                    "status": self._find_value(norm, "acceptance_conclusion"),
                    "source_ref": self._row_source_ref(item, row, index),
                },
                "verification_evidence": self._evidence(self._find_value(norm, "verification_evidence"), item, index),
                "residual_risk": {
                    "summary": self._find_value(norm, "residual_risk"),
                    "source_ref": self._row_source_ref(item, row, index),
                },
                "linked_ifu_refs": [],
                "linked_cer_refs": [],
                "linked_td_refs": [],
                "row_index": index,
                "table_name": row.get("__table_name") or row.get("__sheet_name") or "table_1",
                "raw_cells": [self._clean_cell(cell) for cell in raw_cells],
                "mapping_status": "mapped" if any(value not in (None, "") for value in mapped_fields) else "raw_only",
                "source_ref": self._row_source_ref(item, row, index),
            }
            mapped.append(mapped_row)
        return mapped

    def _infer_rmf_sections(self, rmf_docs: list[dict[str, Any]]) -> list[dict[str, Any]]:
        sections = []
        expected = (
            "risk_management_plan",
            "risk_analysis_matrix",
            "risk_acceptability_matrix",
            "traceability_matrix",
            "residual_risk_evaluation",
            "production_post_production_information",
        )
        for name in expected:
            sections.append(
                {
                    "section_key": name,
                    "status": "candidate_only",
                    "source_refs": [item["source_ref"] for item in rmf_docs],
                }
            )
        return sections

    def _write_run_context(self, *, mode: str) -> None:
        self._write_json(
            self._artifact_path("00_manifest", "runner_context.json"),
            {
                "workflow_path": str(self.workflow_path),
                "project_profile_path": str(self.project_profile_path),
                "mode": mode,
                "thread_id": self.thread_id,
                "run_id": self.run_id,
                "artifact_root_virtual": self.artifact_root_virtual,
                "artifact_root_actual": str(self.artifact_root_actual),
            },
        )

    def _load_prompt_for_step(self, step: dict[str, Any]) -> str:
        prompt_path = (self.repo_root / step["prompt_contract"]).resolve()
        return prompt_path.read_text(encoding="utf-8")

    def _artifact_dir(self, relative: str) -> Path:
        path = self.artifact_root_actual / relative
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _artifact_path(self, directory: str, filename: str) -> Path:
        return self._artifact_dir(directory) / filename

    def _resolve_input_root(self) -> Path:
        if self.input_root_override is not None:
            return self.input_root_override
        root = self.project_profile.get("input_package", {}).get("root_path")
        if not root:
            raise ValueError("project_profile.input_package.root_path is required")
        return Path(root).expanduser().resolve()

    def _resolve_source_path(self, declared_path: str | None, input_root: Path) -> Path:
        if not declared_path:
            return input_root / "__missing__"
        candidate = Path(declared_path).expanduser()
        if candidate.is_absolute():
            return candidate.resolve()
        return (input_root / candidate).resolve()

    def _row_source_ref(self, item: dict[str, Any], row: dict[str, Any], index: int) -> dict[str, Any]:
        return {
            "document_id": item["document_id"],
            "path": item["declared_path"],
            "table_id": row.get("__sheet_name", "table_1"),
            "row_id": str(row.get("__sheet_row", index)),
        }

    def _metric(self, raw_value: Any, item: dict[str, Any]) -> dict[str, Any]:
        return {
            "raw_value": raw_value if raw_value not in (None, "") else None,
            "source_ref": {
                "document_id": item["document_id"],
                "path": item["declared_path"],
            },
        }

    def _controls(self, value: Any, item: dict[str, Any], index: int) -> list[dict[str, Any]]:
        if value in (None, ""):
            return []
        return [
            {
                "control_id": f"{item['document_id']}_control_{index:03d}",
                "description": str(value),
                "source_ref": self._row_source_ref(item, {}, index),
            }
        ]

    def _evidence(self, value: Any, item: dict[str, Any], index: int) -> list[dict[str, Any]]:
        if value in (None, ""):
            return []
        return [
            {
                "evidence_id": f"{item['document_id']}_evidence_{index:03d}",
                "description": str(value),
                "source_ref": self._row_source_ref(item, {}, index),
            }
        ]

    @staticmethod
    def _split_list_value(value: Any) -> list[str]:
        if value in (None, ""):
            return []
        parts = re.split(r"[;\n|]+", str(value))
        return [part.strip() for part in parts if part.strip()]

    @staticmethod
    def _find_header_row(rows: list[tuple[Any, ...]]) -> int | None:
        best_index: int | None = None
        best_score = 0
        for index, row in enumerate(rows[:10]):
            headers = [RMFReviewRunner._clean_cell(value).lower() for value in row if RMFReviewRunner._clean_cell(value)]
            if not headers:
                continue
            score = RMFReviewRunner._score_header_cells(headers)
            if score > best_score:
                best_index = index
                best_score = score
        return best_index if best_score > 0 else None

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _score_header_cells(headers: list[str]) -> int:
        score = 0
        for header in headers:
            normalized = header.lower()
            if any(
                token in normalized
                for token in (
                    "risk",
                    "failure",
                    "effect",
                    "cause",
                    "control",
                    "severity",
                    "occurrence",
                    "失效",
                    "影响",
                    "原因",
                    "控制",
                    "严重度",
                    "发生频度",
                )
            ):
                score += 2
            if any(
                token in normalized
                for token in ("系统", "功能", "零部件", "失效影响", "失效模式", "失效原因", "现行预防控制")
            ):
                score += 3
        return score + len(headers)

    @staticmethod
    def _normalize_key(key: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", key.lower()).strip()

    @staticmethod
    def _find_value(row: dict[str, Any], logical_field: str) -> Any:
        for alias in _COLUMN_ALIASES.get(logical_field, ()):
            alias_key = RMFReviewRunner._normalize_key(alias)
            if alias_key in row and row[alias_key] not in ("", None):
                return row[alias_key]
        return None

    @staticmethod
    def _duplicates(values: list[str]) -> list[str]:
        seen: set[str] = set()
        dupes: set[str] = set()
        for value in values:
            if value in seen:
                dupes.add(value)
            seen.add(value)
        return sorted(dupes)

    def _render_precheck_markdown(self, report: dict[str, Any]) -> str:
        lines = [
            "# FMEA Precheck Report",
            "",
            f"- Row Count: `{report['row_count']}`",
            f"- Structural Status: `{report['structural_status']}`",
            f"- Duplicate Risk IDs: `{len(report['duplicate_risk_ids'])}`",
            f"- Empty Rows: `{report['empty_row_count']}`",
            f"- Orphan Rows: `{report['orphan_row_count']}`",
            f"- Mapping Incomplete Rows: `{report['mapping_incomplete_row_count']}`",
            f"- Manual Review Needed: `{report['manual_review_needed']}`",
            "",
            "## Findings",
            "",
        ]
        if not report["findings"]:
            lines.append("- None")
        else:
            for finding in report["findings"]:
                lines.append(f"- `{finding.get('finding_type', 'finding')}` | risk_id=`{finding.get('risk_id')}`")
        lines.append("")
        return "\n".join(lines)

    def _render_rmf_precheck_markdown(self, report: dict[str, Any]) -> str:
        lines = [
            "# RMF Precheck Report",
            "",
            f"- Structural Status: `{report['structural_status']}`",
            f"- Manual Review Needed: `{report['manual_review_needed']}`",
            f"- Document Count: `{len(report['document_ids'])}`",
            f"- Checks Evaluated: `{len(report['checks'])}`",
            f"- Evidence Hints: `{len(report['evidence_hints'])}`",
            f"- Findings: `{len(report['findings'])}`",
            "",
            "## Checks",
            "",
        ]
        for check in report["checks"]:
            lines.append(f"- `{check['check_key']}` | state=`{check['state']}`")
            for evidence in check.get("evidence", [])[:2]:
                lines.append(f"  evidence: `{evidence['text']}`")
        lines.extend(["", "## Findings", ""])
        if not report["findings"]:
            lines.append("- None")
        else:
            for finding in report["findings"]:
                lines.append(f"- `{finding.get('finding_type', 'finding')}` | `{finding.get('check_key', 'n/a')}` | {finding.get('detail', '')}")
        lines.extend(["", "## Term Consistency", ""])
        lines.append(f"- Status: `{report['term_consistency']['status']}`")
        for note in report["term_consistency"].get("notes", []):
            lines.append(f"- {note}")
        lines.append("")
        return "\n".join(lines)

    def _render_dimension_review_markdown(self, assessment: dict[str, Any]) -> str:
        lines = [
            "# RMF Dimension Review Report",
            "",
            f"- Global Manual Review Needed: `{assessment['global_manual_review_needed']}`",
            f"- RMF Precheck Status: `{assessment['upstream_summary']['rmf_precheck_status']}`",
            f"- FMEA Precheck Status: `{assessment['upstream_summary']['fmea_precheck_status']}`",
            f"- FMEA Row Count: `{assessment['upstream_summary']['fmea_row_count']}`",
            "",
            "## Dimensions",
            "",
        ]
        for dimension_id, dimension in assessment["dimensions"].items():
            lines.append(
                f"- `{dimension_id}` | status=`{dimension['status']}` | human_attention_needed=`{dimension['human_attention_needed']}` | evidence_hints=`{len(dimension['evidence_hints'])}` | findings=`{len(dimension['findings'])}`"
            )
        lines.extend(["", "## Findings Snapshot", ""])
        for dimension_id, dimension in assessment["dimensions"].items():
            if not dimension["findings"]:
                continue
            lines.append(f"### {dimension_id}")
            for finding in dimension["findings"][:3]:
                lines.append(f"- `{finding.get('finding_type', 'finding')}` | {finding.get('detail', '')}")
        lines.append("")
        return "\n".join(lines)

    @staticmethod
    def _search_text_hints(text_lines: list[dict[str, Any]], keywords: list[str]) -> dict[str, Any] | None:
        matches = []
        lowered_keywords = [keyword.lower() for keyword in keywords]
        for entry in text_lines:
            text = entry.get("text", "")
            lowered_text = text.lower()
            if any(keyword in lowered_text for keyword in lowered_keywords):
                matches.append({"text": text[:200], "source_ref": entry.get("source_ref", {})})
        if not matches:
            return None
        return {
            "evidence": matches[:3],
            "source_refs": [match["source_ref"] for match in matches[:3]],
        }

    def _build_term_consistency(
        self,
        *,
        term_map: dict[str, Any],
        text_lines: list[dict[str, Any]],
        cross_doc_entities: dict[str, Any],
        input_inventory: dict[str, Any],
    ) -> dict[str, Any]:
        all_text = "\n".join(entry.get("text", "") for entry in text_lines).lower()
        project_terms = [term["term"] for term in term_map.get("project_terms", []) if term.get("term")]
        canonical_terms = [term["term"] for term in term_map.get("canonical_terms", []) if term.get("term")]
        notes = []
        finding = None

        for project_term in project_terms:
            if project_term.lower() in all_text:
                notes.append(f"Project term `{project_term}` appears in RMF text.")
            else:
                notes.append(f"Project term `{project_term}` was not found in RMF text and needs manual cross-check against IFU.")

        if any(entity.get("doc_type") == "IFU" for entity in cross_doc_entities.get("document_entities", [])):
            notes.append("IFU document is present, so terminology consistency can continue in later review stages.")
        else:
            notes.append("IFU document is absent, so terminology consistency remains weakly evidenced.")

        if not any(term.lower() in all_text for term in canonical_terms):
            finding = {
                "finding_type": "weak_terminology_alignment",
                "check_key": "rmf_ifu_term_consistency",
                "label": "RMF vs IFU key terminology baseline consistency check",
                "severity": "medium",
                "detail": "Canonical RMF/FMEA terminology did not appear clearly in RMF text hints.",
                "source_refs": [entry.get("source_ref", {}) for entry in text_lines[:2]],
            }

        status = "weakly_evidenced" if finding or any("not found" in note for note in notes) else "present"
        return {
            "summary": {
                "status": status,
                "notes": notes,
                "inventory_document_count": len(input_inventory.get("documents", [])),
            },
            "finding": finding,
        }

    def _build_dimension_result(
        self,
        *,
        dimension_id: str,
        label: str,
        evidence_hints: list[dict[str, Any]],
        findings: list[dict[str, Any]],
    ) -> dict[str, Any]:
        status = self._determine_dimension_status(evidence_hints=evidence_hints, findings=findings)
        return {
            "dimension_id": dimension_id,
            "label": label,
            "status": status,
            "evidence_hints": evidence_hints,
            "findings": findings,
            "human_attention_needed": status != "supported",
            "manual_review_needed": True,
        }

    @staticmethod
    def _determine_dimension_status(*, evidence_hints: list[dict[str, Any]], findings: list[dict[str, Any]]) -> str:
        if findings:
            if evidence_hints:
                return "issues_detected"
            return "weakly_evidenced"
        if evidence_hints:
            return "supported"
        return "insufficient_evidence"

    @staticmethod
    def _collect_dimension_evidence(checks_by_key: dict[str, dict[str, Any]], *, keys: list[str]) -> list[dict[str, Any]]:
        evidence: list[dict[str, Any]] = []
        for key in keys:
            check = checks_by_key.get(key)
            if not check:
                continue
            for item in check.get("evidence", [])[:2]:
                evidence.append({"check_key": key, **item})
        return evidence

    @staticmethod
    def _collect_completeness_findings(checks_by_key: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
        findings = []
        for key in ("risk_management_plan", "risk_analysis_matrix", "risk_level_definition", "residual_risk_evaluation"):
            check = checks_by_key.get(key)
            if not check or check.get("state") != "present":
                findings.append(
                    {
                        "finding_type": "coverage_gap",
                        "check_key": key,
                        "detail": f"Completeness evidence for `{key}` is weak or missing.",
                        "source_refs": check.get("source_refs", []) if check else [],
                    }
                )
        return findings

    @staticmethod
    def _collect_correctness_findings(
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
    ) -> list[dict[str, Any]]:
        findings = []
        if fmea_precheck.get("mapping_incomplete_row_count", 0) > 0:
            findings.append(
                {
                    "finding_type": "method_scoring_uncertain",
                    "check_key": "fmea_mapping_quality",
                    "detail": "FMEA mapping is incomplete, so method and scoring correctness require manual review.",
                    "source_refs": [finding.get("source_ref", {}) for finding in fmea_precheck.get("findings", [])[:2]],
                }
            )
        if rmf_precheck.get("term_consistency", {}).get("status") != "present":
            findings.append(
                {
                    "finding_type": "cross_doc_correctness_weak",
                    "check_key": "term_consistency",
                    "detail": "Terminology alignment remains weakly evidenced, so correctness cannot be auto-confirmed.",
                    "source_refs": [],
                }
            )
        return findings

    @staticmethod
    def _collect_adequacy_evidence(fmea_rows: list[dict[str, Any]], checks_by_key: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
        evidence = RMFReviewRunner._collect_dimension_evidence(checks_by_key, keys=["risk_control_hierarchy", "residual_risk_evaluation"])
        for row in fmea_rows[:2]:
            if row.get("controls"):
                evidence.append({"check_key": "fmea_controls", "text": row["controls"][0]["description"], "source_ref": row["controls"][0]["source_ref"]})
        return evidence

    @staticmethod
    def _collect_adequacy_findings(fmea_rows: list[dict[str, Any]], fmea_precheck: dict[str, Any]) -> list[dict[str, Any]]:
        findings = []
        if not any(row.get("controls") for row in fmea_rows):
            findings.append(
                {
                    "finding_type": "control_evidence_weak",
                    "check_key": "control_measures",
                    "detail": "Control measure evidence is weak in normalized FMEA rows.",
                    "source_refs": [],
                }
            )
        if fmea_precheck.get("row_count", 0) and fmea_precheck.get("mapping_incomplete_row_count", 0) > 0:
            findings.append(
                {
                    "finding_type": "adequacy_not_demonstrated",
                    "check_key": "fmea_mapping_quality",
                    "detail": "Some FMEA rows are only partially mapped, so adequacy remains manually reviewable.",
                    "source_refs": [finding.get("source_ref", {}) for finding in fmea_precheck.get("findings", [])[:2]],
                }
            )
        return findings

    @staticmethod
    def _collect_traceability_evidence(
        fmea_rows: list[dict[str, Any]],
        checks_by_key: dict[str, dict[str, Any]],
        cross_doc_entities: dict[str, Any],
    ) -> list[dict[str, Any]]:
        evidence = RMFReviewRunner._collect_dimension_evidence(checks_by_key, keys=["traceability_matrix"])
        for row in fmea_rows[:2]:
            evidence.append({"check_key": "fmea_source_ref", "text": "FMEA row source binding retained.", "source_ref": row.get("source_ref", {})})
        if cross_doc_entities.get("document_entities"):
            evidence.append(
                {
                    "check_key": "cross_doc_entities",
                    "text": f"{len(cross_doc_entities['document_entities'])} document entities available for trace review.",
                    "source_ref": {},
                }
            )
        return evidence

    @staticmethod
    def _collect_traceability_findings(
        fmea_rows: list[dict[str, Any]],
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
    ) -> list[dict[str, Any]]:
        findings = []
        if rmf_precheck.get("term_consistency", {}).get("status") != "present":
            findings.append(
                {
                    "finding_type": "traceability_gap",
                    "check_key": "cross_doc_linkage",
                    "detail": "Cross-document terminology remains weak, so trace linkage cannot be considered complete.",
                    "source_refs": [],
                }
            )
        if any(not row.get("risk_id") for row in fmea_rows[:5]) or fmea_precheck.get("mapping_incomplete_row_count", 0) > 0:
            findings.append(
                {
                    "finding_type": "trace_chain_incomplete",
                    "check_key": "fmea_row_linkage",
                    "detail": "Risk/control/verification chain is incomplete for part of the normalized FMEA bundle.",
                    "source_refs": [finding.get("source_ref", {}) for finding in fmea_precheck.get("findings", [])[:2]],
                }
            )
        return findings

    @staticmethod
    def _collect_consistency_evidence(
        checks_by_key: dict[str, dict[str, Any]],
        inventory_docs: list[dict[str, Any]],
        term_map: dict[str, Any],
    ) -> list[dict[str, Any]]:
        evidence = RMFReviewRunner._collect_dimension_evidence(checks_by_key, keys=["risk_management_plan"])
        if any(doc.get("doc_type") == "IFU" and doc.get("status") == "present" for doc in inventory_docs):
            evidence.append({"check_key": "ifu_presence", "text": "IFU input is present for consistency review.", "source_ref": {}})
        for term in term_map.get("canonical_terms", [])[:2]:
            evidence.append({"check_key": "term_map", "text": term["term"], "source_ref": {}})
        return evidence

    @staticmethod
    def _collect_consistency_findings(
        rmf_precheck: dict[str, Any],
        inventory_docs: list[dict[str, Any]],
        term_map: dict[str, Any],
    ) -> list[dict[str, Any]]:
        findings = []
        if not any(doc.get("doc_type") == "IFU" and doc.get("status") == "present" for doc in inventory_docs):
            findings.append(
                {
                    "finding_type": "ifu_missing",
                    "check_key": "ifu_presence",
                    "detail": "IFU input is absent, so consistency assessment is incomplete.",
                    "source_refs": [],
                }
            )
        if rmf_precheck.get("term_consistency", {}).get("status") != "present":
            findings.append(
                {
                    "finding_type": "term_alignment_weak",
                    "check_key": "term_consistency",
                    "detail": "RMF and IFU terminology alignment remains weakly evidenced.",
                    "source_refs": [],
                }
            )
        if not term_map.get("project_terms"):
            findings.append(
                {
                    "finding_type": "project_term_missing",
                    "check_key": "term_map",
                    "detail": "Project-specific terms are not configured, reducing consistency signal quality.",
                    "source_refs": [],
                }
            )
        return findings

    @staticmethod
    def _collect_acceptability_findings(
        run_manifest: dict[str, Any],
        fmea_precheck: dict[str, Any],
        rmf_precheck: dict[str, Any],
    ) -> list[dict[str, Any]]:
        findings = [
            {
                "finding_type": "human_decision_required",
                "check_key": "residual_risk_acceptability",
                "detail": "Residual risk acceptability requires human judgment and cannot be auto-approved in P0.",
                "source_refs": [],
            }
        ]
        if fmea_precheck.get("mapping_incomplete_row_count", 0) > 0:
            findings.append(
                {
                    "finding_type": "acceptability_evidence_weak",
                    "check_key": "fmea_mapping_quality",
                    "detail": "FMEA mapping gaps weaken acceptability evidence for automated dimension review.",
                    "source_refs": [finding.get("source_ref", {}) for finding in fmea_precheck.get("findings", [])[:2]],
                }
            )
        if run_manifest.get("fmea_status", {}).get("explicit_review_required"):
            findings.append(
                {
                    "finding_type": "human_gate_dependency",
                    "check_key": "explicit_human_gate",
                    "detail": "Workflow requires explicit human gate before any final acceptability conclusion.",
                    "source_refs": [],
                }
            )
        if rmf_precheck.get("structural_status") != "evidence_hints_only":
            findings.append(
                {
                    "finding_type": "acceptability_not_auto_releasable",
                    "check_key": "rmf_precheck_status",
                    "detail": "Upstream RMF precheck still has open issues, so ACPT remains manual-review only.",
                    "source_refs": [],
                }
            )
        return findings

    @staticmethod
    def _determine_rmf_precheck_status(
        *,
        checks: list[dict[str, Any]],
        findings: list[dict[str, Any]],
        evidence_hints: list[dict[str, Any]],
    ) -> str:
        if findings:
            return "issues_detected"
        if evidence_hints or checks:
            return "evidence_hints_only"
        return "no_evidence_extracted"

    @staticmethod
    def _load_yaml(path: Path) -> dict[str, Any]:
        return yaml.safe_load(path.read_text(encoding="utf-8")) or {}

    @staticmethod
    def _read_json(path: Path) -> dict[str, Any]:
        return json.loads(path.read_text(encoding="utf-8"))

    @staticmethod
    def _write_json(path: Path, payload: dict[str, Any] | list[Any]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _write_text(path: Path, content: str) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")

    def _render_template(self, value: str) -> str:
        return value.replace("${run_id}", self.run_id)

    def _apply_hazard_analysis_resolution(self, inventory: list[dict[str, Any]]) -> dict[str, Any]:
        hazard_entries = [item for item in inventory if item.get("doc_type") == "Hazard_Analysis"]
        present_fmea_entries = [item for item in inventory if item.get("doc_type") == "FMEA" and item.get("status") == "present"]
        pfmea_entry = next((item for item in present_fmea_entries if self._looks_like_pfmea(item)), None)
        dfmea_entry = next((item for item in present_fmea_entries if self._looks_like_dfmea(item)), None)

        resolution = {
            "resolution_mode": "none",
            "reason": "No Hazard_Analysis resolution needed.",
        }

        for hazard in hazard_entries:
            if hazard["status"] == "present":
                resolution = {
                    "resolution_mode": "native_present",
                    "reason": "Hazard_Analysis input is present.",
                }
                continue

            if pfmea_entry is not None:
                hazard["status"] = "aliased_to_fmea"
                hazard["blocking_for_p0"] = False
                hazard["alias_target_document_id"] = pfmea_entry["document_id"]
                hazard["alias_target_path"] = pfmea_entry["resolved_path"]
                hazard["resolution_notes"].append("PFMEA used as hazard-related alias input.")
                resolution = {
                    "resolution_mode": "alias_to_pfmea",
                    "reason": "Hazard_Analysis missing; PFMEA present and used as hazard-related alias input.",
                    "alias_target": pfmea_entry["resolved_path"],
                }
            elif dfmea_entry is not None or present_fmea_entries:
                chosen = dfmea_entry or present_fmea_entries[0]
                hazard["status"] = "missing_downgraded"
                hazard["blocking_for_p0"] = False
                hazard["resolution_notes"].append("Hazard_Analysis missing; downgraded because DFMEA/PFMEA is present.")
                resolution = {
                    "resolution_mode": "downgraded_due_to_fmea",
                    "reason": "Hazard_Analysis missing, but DFMEA/PFMEA exists; downgraded to non-blocking for this smoke-run.",
                    "alias_target": chosen["resolved_path"],
                }
            else:
                resolution = {
                    "resolution_mode": "blocking_missing",
                    "reason": "Hazard_Analysis missing and no FMEA-based fallback available.",
                }

        return resolution

    @staticmethod
    def _looks_like_pfmea(item: dict[str, Any]) -> bool:
        haystack = " ".join(str(item.get(field, "")) for field in ("label", "declared_path", "resolved_path")).lower()
        return "pfmea" in haystack

    @staticmethod
    def _looks_like_dfmea(item: dict[str, Any]) -> bool:
        haystack = " ".join(str(item.get(field, "")) for field in ("label", "declared_path", "resolved_path")).lower()
        return "dfmea" in haystack

    @staticmethod
    def _determine_precheck_status(*, rows: list[dict[str, Any]], findings: list[dict[str, Any]]) -> str:
        if not rows:
            return "no_rows_extracted"
        if findings:
            return "issues_detected"
        return "ok"

    @staticmethod
    def _make_run_id() -> str:
        return f"rmf-{uuid.uuid4().hex[:12]}"

    @staticmethod
    def _make_thread_id() -> str:
        return f"rmf-review-{uuid.uuid4().hex[:12]}"
