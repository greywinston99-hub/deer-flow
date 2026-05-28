from deerflow.runtime.cer_authoring.source_preflight import run_source_preflight
from deerflow.runtime.cer_authoring import pipeline
from openpyxl import Workbook


def test_incomplete_working_ifu_is_blocked() -> None:
    text = (
        "等离子手术设备 使用说明书 IFU-JS-00062 A1\n"
        "请企业补充产品型号。请企业补充产品预期用途、预期使用者、预期患者人群。\n"
        "Intended Purpose [blank]. Intended User [blank]. Intended Patient Population [blank].\n"
        "Class IIa under MDR Annex VIII.\n"
    ) * 8
    report = run_source_preflight(
        [
            {
                "source_id": "SRC-IFU",
                "filename": "IFU.docx",
                "document_type": "IFU",
                "source_role": "subject_device_ifu",
                "text": text,
            }
        ],
        {},
    )
    assert report["source_preflight_gate_report"]["status"] == "BLOCKED"
    issue_ids = {row["issue_id"] for row in report["source_preflight_gate_report"]["blocking_issues"]}
    assert "IFU-P0-PLACEHOLDER" in issue_ids


def test_complete_ifu_preflight_passes() -> None:
    text = """
    Plasma Surgical Equipment IFU-JS-00062 A1 Revision A1 effective date 2026-01-01.
    Model ESG301. Intended purpose: used in medical institutions with compatible radiofrequency plasma electrodes for ENT soft tissue cutting and coagulation.
    Intended user: trained ENT surgeons and operating-room clinicians.
    Patient population: patients requiring ENT soft-tissue procedures.
    Clinical benefit: enables controlled cutting and coagulation in saline environment.
    Side effects: bleeding, mucosal injury, infection, pain, thermal injury.
    Contraindications: not for patients where RF/plasma surgery is contraindicated.
    Warnings and precautions: use only with compatible electrodes and saline irrigation.
    Class IIa under MDR Annex VIII.
    """
    report = run_source_preflight(
        [
            {
                "source_id": "SRC-IFU",
                "filename": "IFU.docx",
                "document_type": "IFU",
                "source_role": "subject_device_ifu",
                "text": text,
            }
        ],
        {},
    )
    assert report["source_preflight_gate_report"]["status"] == "PASS"
    assert report["ifu_fact_table"]["status"] == "PASS"


def test_mixed_ifu_domain_is_blocked() -> None:
    plasma = {
        "source_id": "SRC-001",
        "filename": "plasma IFU.docx",
        "document_type": "IFU",
        "source_role": "subject_device_ifu",
        "text": "Plasma Surgical Equipment radiofrequency plasma ENT surgery Class IIa",
    }
    uas = {
        "source_id": "SRC-002",
        "filename": "uas IFU.docx",
        "document_type": "IFU",
        "source_role": "subject_device_ifu",
        "text": "ureteral access sheath renal pelvis ureter Class IIa",
    }
    cardiac = {
        "source_id": "SRC-003",
        "filename": "cardiac IFU.docx",
        "document_type": "IFU",
        "source_role": "subject_device_ifu",
        "text": "Cardiac PFA ablation catheter atrial fibrillation Class III",
    }
    neuro = {
        "source_id": "SRC-004",
        "filename": "neuro IFU.docx",
        "document_type": "IFU",
        "source_role": "subject_device_ifu",
        "text": "Neurovascular stent aneurysm Class III",
    }
    report = run_source_preflight([plasma, uas, cardiac, neuro], {})
    assert report["source_preflight_gate_report"]["status"] == "BLOCKED"
    issue_ids = {row["issue_id"] for row in report["source_preflight_gate_report"]["blocking_issues"]}
    assert "SOURCE-P0-MULTI-IFU" in issue_ids


def test_classification_conflict_is_blocked() -> None:
    text = "Plasma Surgical Equipment IFU. Model ESG301. Class IIa under MDR Annex VIII. Evaluation interval says Class IIb."
    report = run_source_preflight(
        [{"source_id": "SRC-IFU", "filename": "IFU.docx", "document_type": "IFU", "text": text}],
        {},
    )
    assert report["classification_consistency_report"]["status"] == "BLOCKED"


def test_confirmed_manufacturer_intake_resolves_p0_source_blockers(tmp_path) -> None:
    intake = tmp_path / "MANUFACTURER_INTAKE_PACK_TEMPLATE_2026-05-28.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "P0_Device_Scope"
    ws.append(["field_id", "question", "response", "required_level", "status", "note"])
    for row in [
        ("subject_device_boundary", "boundary", "main plasma surgical equipment", "P0", "confirmed", ""),
        ("subject_device_domain", "domain", "plasma_surgical_equipment", "P0", "confirmed", ""),
        ("mdr_classification", "class", "IIb", "P0", "confirmed", ""),
        ("mdr_rule", "rule", "MDR Annex VIII Rule 9", "P0", "confirmed", ""),
        ("classification_rationale", "rationale", "Active therapeutic device transferring RF plasma energy.", "P0", "confirmed", ""),
    ]:
        ws.append(row)
    ws = wb.create_sheet("P0_IFU_Key_Fields")
    ws.append(["field_id", "IFU/CER field", "response", "required_level", "status", "note"])
    for field in [
        "model",
        "intended_purpose",
        "intended_user",
        "patient_population",
        "clinical_benefit",
        "side_effects",
        "warnings",
        "contraindications",
        "document_control",
    ]:
        ws.append([field, field, f"confirmed {field}", "P0", "confirmed", "manufacturer intake"])
    wb.save(intake)

    text = (
        "Plasma Surgical Equipment IFU. 等离子手术设备. 等离子手术电极 accessory reference. "
        "Class IIa under MDR. legacy source also says Class IIb. "
    ) * 8
    report = run_source_preflight(
        [
            {
                "source_id": "SRC-IFU",
                "filename": "IFU.docx",
                "document_type": "IFU",
                "source_role": "subject_device_ifu",
                "text": text,
                "path": str(tmp_path / "IFU.docx"),
            },
            {
                "source_id": "SRC-INTAKE",
                "filename": intake.name,
                "document_type": "manufacturer_intake",
                "text": "",
                "path": str(intake),
            },
        ],
        {},
    )
    assert report["manufacturer_intake_report"]["status"] == "PASS"
    assert report["source_preflight_gate_report"]["status"] != "BLOCKED"
    assert report["classification_consistency_report"]["status"] == "PASS"
    assert report["device_classification_lock"]["device_class"] == "IIb"
    assert report["ifu_fact_table"]["status"] == "PASS"


def test_bubble_study_intake_locks_domain_and_primary_class(tmp_path) -> None:
    intake = tmp_path / "MANUFACTURER_INTAKE_PACK_TEMPLATE_2026-05-28_FILLED.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "P0_Device_Scope"
    ws.append(["field_id", "question", "response", "required_level", "status", "note"])
    for row in [
        ("subject_device_boundary", "boundary", "Bubble Study System with Disposable Contrast Injection Tubing Set", "P0", "confirmed", ""),
        ("subject_device_domain", "domain", "contrast_imaging_bubble_study_system", "P0", "confirmed", ""),
        ("mdr_classification", "class", "Class IIb (system); Class IIa (tubing accessory)", "P0", "confirmed", ""),
        ("mdr_rule", "rule", "Rule 10 + Rule 12; Rule 2 for tubing", "P0", "confirmed", ""),
        ("classification_rationale", "rationale", "Active diagnostic bubble-study system with IV administration risk.", "P0", "confirmed", ""),
    ]:
        ws.append(row)
    ws = wb.create_sheet("P0_IFU_Key_Fields")
    ws.append(["field_id", "IFU/CER field", "response", "required_level", "status", "note"])
    for field in [
        "model",
        "intended_purpose",
        "intended_user",
        "patient_population",
        "clinical_benefit",
        "side_effects",
        "warnings",
        "contraindications",
        "document_control",
    ]:
        ws.append([field, field, f"confirmed {field}", "P0", "confirmed", "manufacturer intake"])
    wb.save(intake)

    report = run_source_preflight(
        [
            {
                "source_id": "SRC-INTAKE",
                "filename": intake.name,
                "document_type": "manufacturer_intake",
                "path": str(intake),
                "text": "",
            }
        ],
        {"input_root": str(tmp_path), "project_id": "WYTD_BUBBLE_STUDY_001"},
    )

    assert report["source_preflight_gate_report"]["status"] == "PASS"
    assert report["source_lock_report"]["locked_domain"] == "contrast_imaging_bubble_study_system"
    assert report["device_classification_lock"]["device_class"] == "IIb"


def test_bubble_study_domain_routes_to_focused_sota_pack() -> None:
    state = {
        "project_id": "WYTD_BUBBLE_STUDY_001",
        "target_keywords": ["bubble study", "agitated saline", "right-to-left shunt", "PFO"],
        "source_lock_report": {"locked_domain": "contrast_imaging_bubble_study_system"},
        "device_profile": {
            "device_name": "Bubble Study System",
            "clinical_domain": "contrast_imaging_bubble_study_system",
            "intended_purpose": "Preparation and injection of agitated saline for c-TTE/c-TCD bubble-study detection of right-to-left shunt and PFO.",
        },
    }

    assert pipeline._clinical_domain(state) == "contrast_imaging_bubble_study_system"
    plan = pipeline._phase7_search_plan(state["device_profile"], state)
    assert plan
    assert {row["retrieval_domain"] for row in plan} == {"contrast_imaging_bubble_study_system"}
    joined_queries = "\n".join(row["query_string"] for row in plan)
    assert "agitated saline" in joined_queries
    assert "patent foramen ovale" in joined_queries
