"""Source preflight checks for CER authoring.

The checks in this module are deterministic and intentionally conservative:
they block only source-package defects that make downstream CER authorship
unsafe, and they surface incomplete manufacturer inputs as explicit gaps.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


PLACEHOLDER_PATTERNS: tuple[tuple[str, str], ...] = (
    ("manufacturer_completion_comment", r"请企业补充|企业补充|请补充"),
    ("blank_marker", r"\[(?:blank|to be completed|date|version|author|reviewer|approver)\]"),
    ("not_extracted", r"\bnot extracted\b|未提取|无法提取"),
    ("pending_confirmation", r"requires confirmation|to be confirmed|待确认|待补充"),
)

REQUIRED_IFU_FIELDS: dict[str, tuple[str, ...]] = {
    "model": ("型号", "model", "specification", "规格"),
    "intended_purpose": ("预期用途", "适用范围", "intended purpose", "intended use", "indication"),
    "intended_user": ("预期使用者", "intended user", "user profile", "operator"),
    "patient_population": ("预期患者", "患者人群", "patient population", "target population"),
    "clinical_benefit": ("临床获益", "clinical benefit", "clinical benefits"),
    "side_effects": ("副作用", "不良反应", "side effect", "adverse effect", "undesirable"),
    "warnings": ("警告", "注意事项", "warning", "precaution"),
    "contraindications": ("禁忌", "contraindication"),
    "document_control": ("ifu-", "版本", "revision", "编写日期", "修改日期", "effective date"),
}

DOMAIN_PATTERNS: dict[str, tuple[str, ...]] = {
    "contrast_imaging_bubble_study_system": (
        "bubble study",
        "agitated saline",
        "contrast imaging",
        "contrast echocardiography",
        "contrast-enhanced transcranial doppler",
        "right-to-left shunt",
        "patent foramen ovale",
        "pfo",
        "rls",
        "c-tte",
        "c-tcd",
        "超声造影",
        "发泡试验",
        "右向左分流",
        "卵圆孔未闭",
        "造影注射",
    ),
    "plasma_surgical_equipment": (
        "等离子手术设备",
        "radiofrequency plasma",
        "plasma surgical equipment",
        "coblation",
        "ENT",
        "耳鼻喉",
    ),
    "plasma_surgical_electrode": ("等离子手术电极", "plasma surgical electrode"),
    "urology_uas": ("输尿管", "肾盂", "access sheath", "ureteral"),
    "cardiac_pfa": ("脉冲电场", "pulsed field", "atrial fibrillation", "心房颤动"),
}

CLASS_PATTERN = re.compile(r"\bclass\s*(IIa|IIb|III|IV|I)\b|(?:Class\s*)?(IIa|IIb|III|IV|I)\s*(?:under MDR|MDR|类)", re.I)


def run_source_preflight(inventory: list[dict[str, Any]], state: dict[str, Any] | None = None) -> dict[str, Any]:
    """Return source lock, IFU fact and preflight gate reports."""
    state = state or {}
    intake_report = _load_manufacturer_intake(inventory, state)
    confirmed_intake = intake_report.get("confirmed_fields") or {}
    intake_domain = _confirmed_value(confirmed_intake, "subject_device_domain")
    intake_class = _normalize_class(_confirmed_value(confirmed_intake, "mdr_classification"))
    ifu_items = _primary_ifu_items(inventory)
    all_text = "\n".join(str(item.get("text") or item.get("extracted_text") or "") for item in inventory)
    ifu_text = "\n".join(str(item.get("text") or item.get("extracted_text") or "") for item in ifu_items)
    intake_text = "\n".join(str(row.get("response") or "") for row in (intake_report.get("rows") or []))
    text_for_domain = f"{all_text}\n{intake_text}\n{state.get('project_id', '')}\n{state.get('input_root', '')}"

    placeholder_hits = _placeholder_hits(ifu_text)
    ifu_fields = _ifu_field_status(ifu_text, confirmed_intake)
    missing_required = [name for name, row in ifu_fields.items() if not row["present"]]
    domain_signals = _domain_signals(text_for_domain)
    class_signals = _classification_signals(all_text)
    primary_domains = [d for d, score in domain_signals.items() if score >= 2]

    blocking_issues: list[dict[str, Any]] = []
    controlled_gaps: list[dict[str, Any]] = []

    if len(ifu_items) > 3:
        blocking_issues.append(
            {
                "issue_id": "SOURCE-P0-MULTI-IFU",
                "severity": "critical",
                "message": "Multiple primary IFU candidates are present; source lock must select one subject-device IFU.",
                "source_ids": [str(item.get("source_id")) for item in ifu_items],
            }
        )

    if len(primary_domains) > 1 and not intake_domain:
        blocking_issues.append(
            {
                "issue_id": "SOURCE-P0-MIXED-DOMAIN",
                "severity": "critical",
                "message": "Strong mixed-device domain signals detected in source package.",
                "domains": primary_domains,
            }
        )
    elif len(primary_domains) > 1 and intake_domain:
        controlled_gaps.append(
            {
                "gap_id": "SOURCE-CONTROLLED-GAP-MIXED-DOMAIN-LOCKED",
                "severity": "major",
                "message": "Mixed source-domain signals remain present, but manufacturer intake supplied a confirmed subject-device domain lock.",
                "domains": primary_domains,
                "locked_domain": intake_domain,
            }
        )

    for hit in placeholder_hits:
        blocking_issues.append(
            {
                "issue_id": "IFU-P0-PLACEHOLDER",
                "severity": "critical",
                "message": f"IFU contains unresolved working-draft placeholder: {hit['kind']}",
                "excerpt": hit["excerpt"],
            }
        )

    if len(set(class_signals)) > 1 and not intake_class:
        blocking_issues.append(
            {
                "issue_id": "CLASS-P0-CONFLICT",
                "severity": "critical",
                "message": "Conflicting MDR classification signals were found.",
                "classes": sorted(set(class_signals)),
            }
        )
    elif len(set(class_signals)) > 1 and intake_class:
        controlled_gaps.append(
            {
                "gap_id": "CLASS-CONTROLLED-GAP-SOURCE-CONFLICT-LOCKED",
                "severity": "major",
                "message": "Conflicting source classification signals remain present, but manufacturer intake supplied a confirmed MDR classification lock.",
                "classes": sorted(set(class_signals)),
                "locked_classification": intake_class,
            }
        )

    if len(ifu_text.strip()) >= 500 and missing_required:
        blocking_issues.append(
            {
                "issue_id": "IFU-P0-REQUIRED-FIELDS-MISSING",
                "severity": "critical",
                "message": "Subject-device IFU is missing required P0 CER authoring fields.",
                "missing_fields": missing_required,
            }
        )
    elif missing_required:
        controlled_gaps.append(
            {
                "gap_id": "IFU-CONTROLLED-GAP-REQUIRED-FIELDS",
                "severity": "major",
                "message": "IFU field completeness could not be fully established from available source text.",
                "missing_fields": missing_required,
            }
        )

    locked_domain = intake_domain or _select_locked_domain(domain_signals, state)
    locked_class = intake_class or (class_signals[0] if len(set(class_signals)) == 1 else "")
    status = "BLOCKED" if blocking_issues else ("REWORK_REQUIRED" if controlled_gaps else "PASS")

    source_lock_report = {
        "schema": "cer_source_lock_report_v1",
        "status": status,
        "created_at": _now(),
        "primary_ifu_source_ids": [str(item.get("source_id")) for item in ifu_items],
        "locked_domain": locked_domain,
        "domain_signals": domain_signals,
        "classification_signals": class_signals,
        "locked_classification": locked_class,
        "blocking_issues": blocking_issues,
        "controlled_gaps": controlled_gaps,
        "manufacturer_intake_status": intake_report.get("status"),
        "manufacturer_intake_path": intake_report.get("path"),
    }
    ifu_fact_table = {
        "schema": "cer_ifu_fact_table_v1",
        "status": "PASS" if not missing_required and not placeholder_hits else "INCOMPLETE",
        "source_ids": [str(item.get("source_id")) for item in ifu_items],
        "fields": ifu_fields,
        "placeholder_hits": placeholder_hits,
    }
    gate_report = {
        "schema": "cer_source_preflight_gate_report_v1",
        "gate_id": "SOURCE_PREFLIGHT",
        "status": status,
        "next_action": "block_authoring_and_emit_report" if status == "BLOCKED" else "continue",
        "blocking_issues": blocking_issues,
        "controlled_gaps": controlled_gaps,
        "manufacturer_intake_status": intake_report.get("status"),
    }
    classification_report = {
        "schema": "cer_classification_consistency_report_v1",
        "status": "BLOCKED" if len(set(class_signals)) > 1 and not intake_class else ("PASS" if locked_class else "CONTROLLED_GAP"),
        "classification_signals": class_signals,
        "locked_classification": locked_class,
        "manufacturer_intake_lock": bool(intake_class),
    }
    classification_lock = {
        "schema": "cer_device_classification_lock_v1",
        "status": "LOCKED" if locked_class else "UNCONFIRMED",
        "device_class": locked_class,
        "source": "manufacturer_intake" if intake_class else "source_preflight",
        "conflicts": sorted(set(class_signals)) if len(set(class_signals)) > 1 and not intake_class else [],
    }
    return {
        "manufacturer_intake_report": intake_report,
        "source_lock_report": source_lock_report,
        "ifu_fact_table": ifu_fact_table,
        "source_preflight_gate_report": gate_report,
        "classification_consistency_report": classification_report,
        "device_classification_lock": classification_lock,
    }


def _primary_ifu_items(inventory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in inventory:
        haystack = " ".join(
            str(item.get(key, ""))
            for key in ("document_type", "doc_type", "type", "filename", "path", "source_role")
        ).lower()
        if "ifu" not in haystack and "使用说明书" not in haystack:
            continue
        if item.get("source_role") in {"similar_device_ifu", "similar_or_benchmark_source", "unconfirmed_ifu"}:
            continue
        if item.get("excluded_from_device_profile"):
            continue
        rows.append(item)
    return rows


def _placeholder_hits(text: str) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    for kind, pattern in PLACEHOLDER_PATTERNS:
        for match in re.finditer(pattern, text or "", re.I):
            start = max(0, match.start() - 60)
            end = min(len(text), match.end() + 80)
            hits.append({"kind": kind, "excerpt": " ".join(text[start:end].split())[:240]})
            if len(hits) >= 20:
                return hits
    return hits


def _ifu_field_status(text: str, confirmed_intake: dict[str, dict[str, Any]] | None = None) -> dict[str, dict[str, Any]]:
    lowered = (text or "").lower()
    confirmed_intake = confirmed_intake or {}
    fields: dict[str, dict[str, Any]] = {}
    for field, tokens in REQUIRED_IFU_FIELDS.items():
        matched = [token for token in tokens if token.lower() in lowered]
        intake_value = _confirmed_value(confirmed_intake, field)
        fields[field] = {
            "present": bool(matched or intake_value),
            "matched_terms": matched[:5],
            "manufacturer_intake_confirmed": bool(intake_value),
        }
    return fields


def _domain_signals(text: str) -> dict[str, int]:
    lowered = (text or "").lower()
    return {
        domain: sum(1 for token in tokens if token.lower() in lowered)
        for domain, tokens in DOMAIN_PATTERNS.items()
    }


def _classification_signals(text: str) -> list[str]:
    found: list[str] = []
    for match in CLASS_PATTERN.finditer(text or ""):
        raw = next((part for part in match.groups() if part), "")
        if not raw:
            continue
        normalized = raw.upper().replace("IIA", "IIa").replace("IIB", "IIb")
        if normalized == "III":
            normalized = "III"
        elif normalized == "IIA":
            normalized = "IIa"
        elif normalized == "IIB":
            normalized = "IIb"
        if normalized not in found:
            found.append(normalized)
    return found[:5]


def _load_manufacturer_intake(inventory: list[dict[str, Any]], state: dict[str, Any]) -> dict[str, Any]:
    paths = _candidate_intake_paths(inventory, state)
    if not paths:
        return {
            "schema": "cer_manufacturer_intake_report_v1",
            "status": "NOT_PROVIDED",
            "path": "",
            "rows": [],
            "confirmed_fields": {},
            "p0_missing_fields": [],
        }
    path = paths[0]
    try:
        rows = _read_intake_workbook(path)
    except Exception as exc:
        return {
            "schema": "cer_manufacturer_intake_report_v1",
            "status": "UNREADABLE",
            "path": str(path),
            "error": f"{type(exc).__name__}: {exc}",
            "rows": [],
            "confirmed_fields": {},
            "p0_missing_fields": [],
        }
    confirmed = {
        str(row.get("field_id") or row.get("control_id") or ""): row
        for row in rows
        if _is_confirmed_status(row.get("status")) and str(row.get("response") or "").strip()
    }
    p0_rows = [
        row for row in rows
        if str(row.get("required_level") or "").strip().upper() == "P0"
    ]
    p0_missing = [
        str(row.get("field_id") or row.get("control_id") or "")
        for row in p0_rows
        if not (_is_confirmed_status(row.get("status")) and str(row.get("response") or "").strip())
    ]
    return {
        "schema": "cer_manufacturer_intake_report_v1",
        "status": "PASS" if not p0_missing else "INCOMPLETE",
        "path": str(path),
        "rows": rows,
        "confirmed_fields": confirmed,
        "p0_missing_fields": p0_missing,
        "p0_total": len(p0_rows),
        "p0_confirmed": len(p0_rows) - len(p0_missing),
    }


def _candidate_intake_paths(inventory: list[dict[str, Any]], state: dict[str, Any]) -> list[Path]:
    candidates: list[Path] = []
    for item in inventory:
        raw = str(item.get("path") or item.get("file_path") or item.get("filename") or "")
        if not raw:
            continue
        lowered = raw.lower()
        if "manufacturer_intake" in lowered and lowered.endswith(".xlsx"):
            candidates.append(Path(raw).expanduser())
    input_root = state.get("input_root")
    if input_root:
        root = Path(str(input_root)).expanduser()
        candidates.extend(sorted((root / "00_MANUFACTURER_INTAKE").glob("*manufacturer_intake*.xlsx")))
        candidates.extend(sorted((root.parent / "00_MANUFACTURER_INTAKE").glob("*manufacturer_intake*.xlsx")))
    existing = []
    seen = set()
    for path in candidates:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        key = str(resolved)
        if key in seen or not resolved.exists():
            continue
        seen.add(key)
        existing.append(resolved)
    return existing


def _read_intake_workbook(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in ("P0_Device_Scope", "P0_IFU_Key_Fields", "P1_Evidence_Controls", "P2_Doc_Control"):
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        header_row = None
        headers: list[str] = []
        for row_idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
            normalized = [_normalize_header(value) for value in values]
            if "field_id" in normalized or "control_id" in normalized:
                header_row = row_idx
                headers = normalized
                break
        if header_row is None:
            continue
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row = {headers[idx]: _cell_text(values[idx]) for idx in range(min(len(headers), len(values))) if headers[idx]}
            field_id = row.get("field_id") or row.get("control_id")
            if not field_id:
                continue
            row["sheet"] = sheet_name
            rows.append(row)
    return rows


def _normalize_header(value: Any) -> str:
    header = str(value or "").strip().lower()
    header = re.sub(r"\s*/\s*", "_", header)
    header = re.sub(r"[^a-z0-9_]+", "_", header)
    header = re.sub(r"_+", "_", header).strip("_")
    if header in {"response_summary", "response"}:
        return "response"
    if header in {"file_or_table_reference", "source_location_or_note", "evidence_source_or_note", "note"}:
        return "note"
    return header


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_confirmed_status(status: Any) -> bool:
    return str(status or "").strip().lower() in {"confirmed", "approved", "final", "locked"}


def _confirmed_value(confirmed: dict[str, dict[str, Any]], field_id: str) -> str:
    row = confirmed.get(field_id) or {}
    return str(row.get("response") or "").strip()


def _normalize_class(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    matches = []
    for match in CLASS_PATTERN.finditer(raw):
        part = next((item for item in match.groups() if item), "")
        if part:
            matches.append(part)
    if not matches:
        for pattern in ("IV", "III", "IIb", "IIa", "I"):
            if re.search(rf"\b{re.escape(pattern)}\b", raw, re.I):
                matches.append(pattern)
    if matches:
        normalized_matches = [_normalize_class_token(item) for item in matches]
        ranking = {"I": 1, "IIa": 2, "IIb": 3, "III": 4, "IV": 5}
        return max(normalized_matches, key=lambda item: ranking.get(item, 0))
    compact = raw.replace("Class", "").replace("class", "").strip()
    lowered = compact.lower()
    if lowered in {"i", "class i"}:
        return "I"
    if lowered in {"iia", "ii a", "class iia"}:
        return "IIa"
    if lowered in {"iib", "ii b", "class iib"}:
        return "IIb"
    if lowered in {"iii", "class iii"}:
        return "III"
    return raw


def _normalize_class_token(value: str) -> str:
    compact = str(value or "").replace("Class", "").replace("class", "").strip()
    lowered = compact.lower().replace(" ", "")
    if lowered == "i":
        return "I"
    if lowered == "iia":
        return "IIa"
    if lowered == "iib":
        return "IIb"
    if lowered == "iii":
        return "III"
    if lowered == "iv":
        return "IV"
    return compact


def _select_locked_domain(signals: dict[str, int], state: dict[str, Any]) -> str:
    hinted = str(state.get("device_domain") or state.get("clinical_domain") or "")
    if hinted:
        return hinted
    if not signals:
        return ""
    domain, score = max(signals.items(), key=lambda item: item[1])
    return domain if score > 0 else ""


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()
