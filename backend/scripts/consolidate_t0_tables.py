#!/usr/bin/env python3
"""将 CER 输出目录中散落的 xlsx/json 文件按 CER 章节合并为 T0 聚合工作簿。

用法:
    python3 consolidate_t0_tables.py <artifact_root> [--output <output_path>]

输出:
    T0_CER_aggregated.xlsx  — 多 sheet 工作簿，每个 sheet 对应一个 CER 章节
    intermediate/           — T1/T2 中间产物移动到此目录
"""

import json
import os
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ── T0 mapping: CER chapter → source files ──
T0_MAPPING = {
    "02_Scope_Device": [
        "source_inventory.xlsx",
        "device_profile.json",
    ],
    "02_Scope_Claims": [
        "claim_ledger.xlsx",
        "claim_pico_derivation.xlsx",
        "intended_purpose_claim_table",  # embedded in workbook
    ],
    "03_SOTA_Search": [
        "sota_search_strategy_table.xlsx",
        "sota_search_strategy_separated.xlsx",
        "database_search_source_table.xlsx",
        "search_protocol_and_results.docx",  # skip binary
    ],
    "03_SOTA_Benchmarks": [
        "sota_benchmark_table.xlsx",
        "sota_benchmark_matrix.xlsx",
        "sota_clinical_context_table.xlsx",
        "sota_benchmark_contextual_rationale.xlsx",
        "sota_conclusion_strength_guard.xlsx",
    ],
    "03_SOTA_Alternatives": [
        "alternative_treatment_benchmark_table.xlsx",
        "guideline_pathway_table.xlsx",
        "similar_benchmark_device_table.xlsx",
        "hazard_source_table.xlsx",
    ],
    "04_Evidence_Screening": [
        "screening_disposition_table.xlsx",
        "sota_screening_disposition_table.xlsx",
        "literature_flow_registry.xlsx",
        "literature_defined_limits.xlsx",
        "prisma_flow_data.json",
    ],
    "04_Evidence_Appraisal": [
        "evidence_appraisal_table.xlsx",
        "evidence_source_inventory.xlsx",
        "due_suitability_contribution_table.xlsx",
        "sota_ck_appraisal_table.xlsx",
        "evidence_funnel_counts.json",
    ],
    "04_Evidence_Facts": [
        "clinical_evidence_fact_table.xlsx",
        "endpoint_extraction_table.xlsx",
        "endpoint_registry.xlsx",
    ],
    "04_Evidence_ClaimMatrix": [
        "claim_evidence_matrix.xlsx",
        "pre_g42_claim_evidence_candidate_matrix.xlsx",
        "semantic_claim_evidence_candidate_matrix.xlsx",
        "claim_support_type_classifier.xlsx",
        "claim_support_matrix.json",
        "final_text_claim_support_map.json",
    ],
    "04_Evidence_Synthesis": [
        "cross_evidence_synthesis_table.xlsx",
        "cross_evidence_synthesis_narratives.xlsx",
        "evidence_conflict_report.json",
    ],
    "04_Equivalence": [
        "equivalence_comparison_matrix.xlsx",
        "equivalence_3d_comparison_table.xlsx",
        "similar_device_attachment_index.xlsx",
        "similar_device_four_step_confirmation.xlsx",
    ],
    "04_Vigilance": [
        "vigilance_event_statistics.xlsx",
        "vigilance_recall_registry.xlsx",
    ],
    "04_BenefitRisk": [
        "benefit_risk_ledger.xlsx",
        "benefit_risk_conclusion.json",
    ],
    "04_Risk_GSPR": [
        "risk_gspr_trace_matrix.xlsx",
    ],
    "04_PMCF": [
        "pmcf_gap_register.xlsx",
        "pmcf_boundary_decision_log.xlsx",
        "gap_pmcf_recommendations.docx",  # skip binary
    ],
}

# ── Files that should stay in output root (final deliverables) ──
KEEP_IN_ROOT = {
    "T0_CER_aggregated.xlsx",
    "CER_draft.docx",
    "CER_draft.md",
    "nb_precheck_report.docx",
    "search_protocol_and_results.docx",
    "gap_pmcf_recommendations.docx",
    "cer_dashboard.html",
    "device_profile.json",
    "authoring_workbook.json",
    "FINAL_DRAFT_QA_REPORT.json",
    "final_gate_closure_report.json",
    "qa_gate_report.json",
    "context_contamination_trace.json",
    "MODEL_RESOLUTION_TRACE.json",
}

INTERMEDIATE_DIR = "intermediate"
CLAUDE_TEAM_DIR = "claude_team"


def style_header(ws, num_cols):
    """Apply header styling."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border


def auto_width(ws, max_width=60):
    """Auto-fit column widths."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def _safe_cell_value(val):
    """Convert any Python value to an Excel-compatible string."""
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        return json.dumps(val, ensure_ascii=False)[:2000]
    return str(val)[:2000]


def json_to_rows(json_path: str) -> tuple[list[str], list[list]]:
    """Convert a JSON file to table rows. Returns (headers, rows)."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        # Try to find a list field first
        for key in data:
            if isinstance(data[key], list) and data[key]:
                items = data[key]
                if isinstance(items[0], dict):
                    headers = list(items[0].keys())
                    rows = [[_safe_cell_value(item.get(h)) for h in headers] for item in items]
                    return headers, rows
        # Fall back to key-value pairs
        return ["Key", "Value"], [[str(k), _safe_cell_value(v)] for k, v in data.items()]

    if isinstance(data, list) and data and isinstance(data[0], dict):
        headers = list(data[0].keys())
        rows = [[_safe_cell_value(item.get(h)) for h in headers] for item in data]
        return headers, rows

    return ["Value"], [[_safe_cell_value(data)]]


def xlsx_to_rows(xlsx_path: str) -> list[tuple[str, list[str], list[list]]]:
    """Convert an xlsx file to sheets. Returns [(sheet_name, headers, rows), ...]."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return []
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_safe_cell_value(c) for c in rows[0]]
        data = [[_safe_cell_value(c) for c in row] for row in rows[1:]]
        result.append((sheet_name, headers, data))
    wb.close()
    return result


def consolidate(artifact_root: str, output_path: str):
    """Main consolidation logic."""
    root = Path(artifact_root)
    if not root.exists():
        print(f"ERROR: Artifact root not found: {artifact_root}")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    summary_rows = [["Sheet", "Source Files", "Row Count"]]

    for chapter, sources in T0_MAPPING.items():
        sheet_data = []
        source_names = []
        for src in sources:
            src_path = root / src
            if not src_path.exists():
                continue
            source_names.append(src)
            if src.endswith(".json"):
                try:
                    headers, rows = json_to_rows(str(src_path))
                    if rows:
                        sheet_data.append((f"{src}", headers, rows))
                except Exception:
                    continue
            elif src.endswith(".xlsx"):
                sheets = xlsx_to_rows(str(src_path))
                for sh_name, headers, rows in sheets:
                    if rows:
                        label = f"{src}/{sh_name}" if sh_name != "Sheet1" else src
                        sheet_data.append((label, headers, rows))

        if not sheet_data:
            continue

        # Write to a single sheet per chapter
        ws = wb.create_sheet(title=chapter[:31])  # Excel sheet name max 31 chars
        row_idx = 1
        for label, headers, data_rows in sheet_data:
            # Source label row
            ws.cell(row=row_idx, column=1, value=f"Source: {label}")
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="1F4E79")
            row_idx += 1
            # Headers
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=h)
            style_header_single_row(ws, row_idx, len(headers))
            row_idx += 1
            # Data
            for data_row in data_rows:
                for col_idx, val in enumerate(data_row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                row_idx += 1
            row_idx += 1  # Blank row between sources

        summary_rows.append([chapter, "; ".join(source_names), str(row_idx - 1)])

    # Summary sheet
    ws_summary = wb.create_sheet(title="00_Summary", index=0)
    for row_idx, row in enumerate(summary_rows, start=1):
        for col_idx, val in enumerate(row, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=val)
    style_header(ws_summary, 3)
    auto_width(ws_summary)

    out = Path(output_path)
    wb.save(str(out))
    print(f"T0 aggregated workbook saved to: {out}")


def style_header_single_row(ws, row_idx, num_cols):
    """Style a single header row."""
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def organize_directories(artifact_root: str):
    """Move intermediate files to intermediate/ and claude_team/ directories."""
    root = Path(artifact_root)
    intermediate = root / INTERMEDIATE_DIR
    claude_team = root / CLAUDE_TEAM_DIR
    intermediate.mkdir(exist_ok=True)
    claude_team.mkdir(exist_ok=True)

    # All xlsx files not in T0_MAPPING sources go to intermediate
    t0_sources = set()
    for sources in T0_MAPPING.values():
        for s in sources:
            t0_sources.add(s)

    moved_intermediate = []
    moved_claude_team = []

    for f in root.iterdir():
        if f.is_dir():
            if f.name in (INTERMEDIATE_DIR, CLAUDE_TEAM_DIR, ".cache", ".human_gate"):
                continue
            continue
        if f.name in KEEP_IN_ROOT:
            continue
        if f.suffix in (".xlsx", ".csv"):
            if f.name in t0_sources:
                continue
            shutil.move(str(f), str(intermediate / f.name))
            moved_intermediate.append(f.name)
        elif f.suffix in (".json", ".md") and f.name not in KEEP_IN_ROOT:
            if f.name.startswith("calibration") or f.name.startswith("gate_routing") or \
               f.name.startswith("artifact_consumption") or f.name.startswith("failure_taxonomy") or \
               f.name.startswith("cer_section_trace_map_schema"):
                shutil.move(str(f), str(claude_team / f.name))
                moved_claude_team.append(f.name)
            else:
                shutil.move(str(f), str(intermediate / f.name))
                moved_intermediate.append(f.name)

    if moved_intermediate:
        print(f"Moved {len(moved_intermediate)} files to {INTERMEDIATE_DIR}/")
    if moved_claude_team:
        print(f"Moved {len(moved_claude_team)} files to {CLAUDE_TEAM_DIR}/")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    artifact_root = sys.argv[1]
    output_path = sys.argv[3] if len(sys.argv) > 3 and sys.argv[2] == "--output" else \
        os.path.join(artifact_root, "T0_CER_aggregated.xlsx")

    print(f"Consolidating T0 tables from: {artifact_root}")
    consolidate(artifact_root, output_path)
    organize_directories(artifact_root)
    print("Done.")


if __name__ == "__main__":
    main()
